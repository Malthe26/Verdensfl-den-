# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
import shutil
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill


# ============================================================
# CONFIGURATION - EDIT THESE PATHS
# ============================================================

INPUT_FOLDER = Path(r"C:\Users\mrt\OneDrive - Danske Rederier\Desktop\Verdensflaaden\Juli 2026 (30-06-2026)")

# Final workbook. Keep this outside INPUT_FOLDER if possible.
OUT_PATH = Path(r"W:\Data\Verdensflåden\2026\Verdensflåden.2026.06.31.xlsx")

PREFERRED_INPUT_SHEET = "Sheet"
OUTPUT_SHEET_NAME = "Sheet1"

# Use calamine for faster input reading. Output is still written with openpyxl
# because the script formats the final workbook with openpyxl styles.
READ_ENGINE = "calamine"
FALLBACK_READ_ENGINE = "openpyxl"

# Auto-select workers based on the PC, but keep it within a safe laptop range.
# Example: 8 logical CPU threads -> 6 workers.
MAX_READ_WORKERS = min(8, max(4, (os.cpu_count() or 4) - 2))

# The final workbook must have exactly these columns and this order.
OUTPUT_COLUMNS = [
    "Source.Name",
    "IMO/LR/IHS No.",
    "Name of Ship",
    "Flag",
    "GT",
    "Deadweight",
    "TEU",
    "Ship Type",
    "Operator Domicile",
    "Operator",
    "Group Owner Domicile",
    "Group Owner",
    "Build Location",
    "Build Year",
]

# Segment naming used only when the input is still in subfolders.
SEGMENT_ORDER = ["Bulk", "GC", "Inland", "Misc", "Offshore", "Tankers"]
SEGMENT_OUTPUT_PREFIX = {
    "bulk": "Bulk",
    "gc": "GC",
    "inland": "Inland",
    "misc": "MISC",
    "offshore": "Offshore",
    "tankers": "TANKERS",
}

# The March workbook has Inland.xlsx, not Inland 1.xlsx, because it is one file.
SINGLE_FILE_WITHOUT_NUMBER = {"inland"}

# If a monthly backup still has files sorted into segment subfolders
# (Bulk\, GC\, Inland\, Misc\, Offshore\, Tankers\), move them up into
# INPUT_FOLDER and rename them to the flat naming convention before reading,
# so every monthly backup ends up looking the same on disk.
FLATTEN_SEGMENT_FOLDERS = True


# ============================================================
# BASIC HELPERS
# ============================================================

def natural_key(value: str):
    """Sort strings naturally, e.g. Bulk 2 before Bulk 10."""
    return tuple(int(x) if x.isdigit() else x.lower() for x in re.split(r"(\d+)", str(value)))


def clean_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [clean_header(c) for c in df.columns]

    rename_map = {
        "Source Name": "Source.Name",
        "Sourcename": "Source.Name",
        "Source": "Source.Name",
        "IMO/LR/HS No.": "IMO/LR/IHS No.",
        "IMO LR IHS No.": "IMO/LR/IHS No.",
        "IMO No.": "IMO/LR/IHS No.",
        "IMO": "IMO/LR/IHS No.",
        "Built": "Build Year",
        "Build Date": "Build Year",
    }
    df.rename(columns={c: rename_map.get(c, c) for c in df.columns}, inplace=True)
    return df


def normalize_country_name(value: object) -> str:
    """
    Normalise country names.

    Important:
    Denmark, Denmark (DIS), Denmark (DAS), etc. are treated as Denmark.
    Same logic also handles Norway (NIS), etc.
    """
    if value is None or pd.isna(value):
        return ""

    text = str(value).replace("\xa0", " ").strip()
    if not text or text.lower() in {"nan", "none"}:
        return ""

    text = re.sub(r"\s*\((dis|nis|mar|fis|fas|das)\)\s*$", "", text, flags=re.IGNORECASE)
    return text.strip()


def remove_patrol_vessels(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Remove patrol vessels because naval/warship tonnage is out of scope."""
    if "Ship Type" not in df.columns:
        return df, 0

    ship_type = (
        df["Ship Type"]
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .str.lower()
    )

    mask_patrol = ship_type.str.contains(r"\bpatrol\s+vessels?\b", regex=True, na=False)
    removed = int(mask_patrol.sum())
    return df.loc[~mask_patrol].copy(), removed


def read_excel_file(path: Path) -> pd.DataFrame:
    """
    Read one Seaweb/IHS file using row 2 as header, matching the Power Query logic.

    Tries calamine first because it is faster for reading .xlsx files.
    Falls back to openpyxl if calamine is not installed or one workbook cannot be read.
    """
    try:
        with pd.ExcelFile(path, engine=READ_ENGINE) as xls:
            sheet_name = PREFERRED_INPUT_SHEET if PREFERRED_INPUT_SHEET in xls.sheet_names else xls.sheet_names[0]
            df = pd.read_excel(xls, sheet_name=sheet_name, header=1, dtype=object)
    except Exception as exc:
        print(f"  Calamine failed for {path.name}; using openpyxl instead. Reason: {exc}")
        with pd.ExcelFile(path, engine=FALLBACK_READ_ENGINE) as xls:
            sheet_name = PREFERRED_INPUT_SHEET if PREFERRED_INPUT_SHEET in xls.sheet_names else xls.sheet_names[0]
            df = pd.read_excel(xls, sheet_name=sheet_name, header=1, dtype=object)

    df = normalize_columns(df)
    df.dropna(how="all", inplace=True)
    return df


def read_source_frame(file_path: Path, source_name: str) -> pd.DataFrame:
    df = read_excel_file(file_path)
    df.drop(columns=["Source.Name", "Source Name"], errors="ignore", inplace=True)
    df.insert(0, "Source.Name", source_name)
    return df


def flatten_segment_folders(input_folder: Path) -> None:
    """
    Move files out of segment subfolders (Bulk\\, GC\\, Inland\\, Misc\\,
    Offshore\\, Tankers\\) into input_folder and rename them to the flat
    naming convention (Bulk 1.xlsx, GC 1.xlsx, MISC 1.xlsx, Inland 1.xlsx, ...),
    matching how newer monthly backups already look. Empty subfolders are
    removed afterwards.

    Does nothing if input_folder has no recognised segment subfolders.
    """
    if not input_folder.exists():
        return

    segment_folders = [
        p for p in input_folder.iterdir()
        if p.is_dir() and p.name.strip().lower() in SEGMENT_OUTPUT_PREFIX
    ]
    if not segment_folders:
        return

    print(f"Flatter {len(segment_folders)} segmentmappe(r) i: {input_folder}")

    for folder in sorted(segment_folders, key=lambda p: natural_key(p.name)):
        segment_key = folder.name.strip().lower()
        prefix = SEGMENT_OUTPUT_PREFIX.get(segment_key, folder.name.strip().title())

        files = sorted(
            [f for f in folder.glob("*.xlsx") if not f.name.startswith("~$")],
            key=lambda p: natural_key(p.name),
        )

        for i, file_path in enumerate(files, start=1):
            if len(files) == 1 and segment_key in SINGLE_FILE_WITHOUT_NUMBER:
                new_name = f"{prefix}.xlsx"
            else:
                new_name = f"{prefix} {i}.xlsx"

            dest_path = input_folder / new_name
            if dest_path.exists():
                print(f"  Springer over (findes allerede i {input_folder.name}): {new_name}")
                continue

            shutil.move(str(file_path), str(dest_path))
            print(f"  {folder.name}/{file_path.name} -> {new_name}")

        remaining = list(folder.glob("*"))
        if not remaining:
            try:
                shutil.rmtree(str(folder))
                print(f"  Fjernede tom mappe: {folder.name}")
            except (OSError, PermissionError) as e:
                print(f"  ⚠️  Kunne ikke fjerne mappe (er måske låst): {folder.name} - {e}")
        else:
            print(f"  Behold mappe (indeholder stadig filer): {folder.name}")


def detect_input_files(input_folder: Path) -> list[tuple[Path, str]]:
    """
    Return list of (file_path, source_name_for_output).

    If files are in segment subfolders, generate names like Bulk 1.xlsx, GC 1.xlsx,
    MISC 1.xlsx, TANKERS 1.xlsx, etc.
    If no segment subfolders exist, direct .xlsx files in INPUT_FOLDER are used.
    """
    if not input_folder.exists():
        sys.exit(f"INPUT_FOLDER does not exist: {input_folder}")

    if FLATTEN_SEGMENT_FOLDERS:
        flatten_segment_folders(input_folder)

    files_by_segment: dict[str, list[Path]] = defaultdict(list)

    for folder in sorted([p for p in input_folder.iterdir() if p.is_dir()], key=lambda p: natural_key(p.name)):
        segment_key = folder.name.strip().lower()
        for f in sorted(folder.glob("*.xlsx"), key=lambda p: natural_key(p.name)):
            if not f.name.startswith("~$"):
                files_by_segment[segment_key].append(f)

    if not files_by_segment:
        direct_files = sorted(
            [f for f in input_folder.glob("*.xlsx") if not f.name.startswith("~$")],
            key=lambda p: natural_key(p.name),
        )
        if direct_files:
            return [(f, f.name) for f in direct_files if f.resolve() != OUT_PATH.resolve()]
        sys.exit(f"No .xlsx files found in: {input_folder}")

    ordered_segment_keys: list[str] = []
    for seg in SEGMENT_ORDER:
        key = seg.lower()
        if key in files_by_segment:
            ordered_segment_keys.append(key)

    for key in sorted(files_by_segment.keys(), key=natural_key):
        if key not in ordered_segment_keys:
            ordered_segment_keys.append(key)

    result: list[tuple[Path, str]] = []
    for segment_key in ordered_segment_keys:
        files = files_by_segment[segment_key]
        prefix = SEGMENT_OUTPUT_PREFIX.get(segment_key, segment_key.title())

        for i, file_path in enumerate(files, start=1):
            if len(files) == 1 and segment_key in SINGLE_FILE_WITHOUT_NUMBER:
                source_name = f"{prefix}.xlsx"
            else:
                source_name = f"{prefix} {i}.xlsx"
            result.append((file_path, source_name))

    return result


def coerce_final_types(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "IMO/LR/IHS No." in df.columns:
        df["IMO/LR/IHS No."] = df["IMO/LR/IHS No."].astype(str).str.strip()
        df = df[df["IMO/LR/IHS No."].ne("") & df["IMO/LR/IHS No."].ne("nan")].copy()

    for col in ["GT", "Deadweight", "TEU"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            df[col] = df[col].where(df[col].isna(), df[col].round(0).astype("Int64"))

    if "Build Year" in df.columns:
        df["Build Year"] = df["Build Year"].astype(str).replace({"nan": ""})

    return df


def format_output_worksheet(ws) -> None:
    """Format the final world fleet sheet."""
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    widths = {
        "A": 13,
        "B": 13,
        "C": 49,
        "D": 30,
        "E": 13,
        "F": 13,
        "G": 13,
        "H": 13,
        "I": 25,
        "J": 13,
        "K": 13,
        "L": 13,
        "M": 13,
        "N": 13,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = "@"
        ws.cell(row=row, column=14).number_format = "@"

    ws.auto_filter.ref = None
    ws.freeze_panes = None


# ============================================================
# MAIN FLOW
# ============================================================

def read_all_input_frames(file_map: list[tuple[Path, str]]) -> tuple[list[pd.DataFrame], list[tuple[str, str]]]:
    frames: list[pd.DataFrame] = []
    failed_files: list[tuple[str, str]] = []

    worker_count = max(1, min(MAX_READ_WORKERS, len(file_map)))
    print(f"Reading input files: {len(file_map)} file(s), {worker_count} worker(s)")

    if worker_count == 1:
        for file_path, source_name in file_map:
            try:
                frames.append(read_source_frame(file_path, source_name))
            except Exception as exc:
                failed_files.append((str(file_path), str(exc)))
        return frames, failed_files

    ordered_frames: list[pd.DataFrame | None] = [None] * len(file_map)
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(read_source_frame, file_path, source_name): (index, file_path)
            for index, (file_path, source_name) in enumerate(file_map)
        }

        for future in as_completed(futures):
            index, file_path = futures[future]
            try:
                ordered_frames[index] = future.result()
            except Exception as exc:
                failed_files.append((str(file_path), str(exc)))

    frames = [frame for frame in ordered_frames if frame is not None]
    return frames, failed_files


def build_world_fleet(file_map: list[tuple[Path, str]]) -> tuple[pd.DataFrame, dict[str, int], list[tuple[str, str]]]:
    frames, failed_files = read_all_input_frames(file_map)

    if not frames:
        sys.exit("No readable input files. Stopping.")

    combined = pd.concat(frames, ignore_index=True)
    combined = normalize_columns(combined)

    rows_before_scope_filter = len(combined)
    combined, patrol_removed = remove_patrol_vessels(combined)
    rows_after_scope_filter = len(combined)

    for col in OUTPUT_COLUMNS:
        if col not in combined.columns:
            combined[col] = ""

    combined = combined[OUTPUT_COLUMNS]
    combined = coerce_final_types(combined)

    rows_before_dedup = len(combined)
    if "IMO/LR/IHS No." in combined.columns:
        combined = combined.drop_duplicates(subset=["IMO/LR/IHS No."], keep="first").copy()
    rows_after_dedup = len(combined)

    stats = {
        "rows_before_scope_filter": rows_before_scope_filter,
        "patrol_removed": patrol_removed,
        "rows_after_scope_filter": rows_after_scope_filter,
        "rows_before_dedup": rows_before_dedup,
        "rows_after_dedup": rows_after_dedup,
    }

    return combined, stats, failed_files


def write_world_fleet_output(combined: pd.DataFrame) -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    if OUT_PATH.exists():
        OUT_PATH.unlink()

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME, index=False)
        format_output_worksheet(writer.sheets[OUTPUT_SHEET_NAME])


def main() -> None:
    total_start = time.perf_counter()

    detect_start = time.perf_counter()
    file_map = detect_input_files(INPUT_FOLDER)
    print(f"Detect input files: {time.perf_counter() - detect_start:.2f} sec")

    build_start = time.perf_counter()
    combined, stats, failed_files = build_world_fleet(file_map)
    print(f"Read/build data: {time.perf_counter() - build_start:.2f} sec")

    write_start = time.perf_counter()
    write_world_fleet_output(combined)
    print(f"Write output: {time.perf_counter() - write_start:.2f} sec")

    print("Done")
    print(f"Total runtime: {time.perf_counter() - total_start:.2f} sec")
    print(f"Input files read: {len(file_map)}")
    print(f"Rows before Patrol Vessel filter: {stats['rows_before_scope_filter']}")
    print(f"Patrol Vessel rows removed: {stats['patrol_removed']}")
    print(f"Rows after Patrol Vessel filter: {stats['rows_after_scope_filter']}")
    print(f"Rows before IMO deduplication: {stats['rows_before_dedup']}")
    print(f"IMO duplicates removed: {stats['rows_before_dedup'] - stats['rows_after_dedup']}")
    print(f"Final rows: {stats['rows_after_dedup']}")
    print(f"Output workbook: {OUT_PATH}")

    if failed_files:
        print("Files skipped due to errors:")
        for file_name, error in failed_files:
            print(f"- {file_name}: {error}")


if __name__ == "__main__":
    main()
