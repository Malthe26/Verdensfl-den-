from __future__ import annotations

import re
import shutil
import sys
import os
import stat
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

# ============================================================
# CONFIGURATION - EDIT THESE PATHS
# ============================================================

INPUT_FOLDER = Path(r"C:\Users\mrt\OneDrive - Danske Rederier\Desktop\Verdensflåden\2025.12.05")

# Final workbook. Keep this outside INPUT_FOLDER if possible.
OUT_PATH = Path(r"W:\Data\Verdensflaaden\2026\Verdensflaaden.2026.03.24.xlsx")

# Optional preview workbook showing the exact fact-sheet splits before writing to WorkPoint.
CREATE_FACTAARK_PREVIEW = True
FACTAARK_PREVIEW_PATH = OUT_PATH.with_name(f"{OUT_PATH.stem}_faktaark_preview.xlsx")

# If the workbook has a sheet called Sheet, use it. Otherwise the first sheet is used.
PREFERRED_INPUT_SHEET = "Sheet"
OUTPUT_SHEET_NAME = "Sheet1"

# The final workbook must have exactly these columns and this order.
OUTPUT_COLUMNS = [
    "Source.Name",
    "IMO/LR/IHS No.",
    "Name of Ship",
    "Flag",
    "GT",
    "Deadweight",
    "TEU",
    "Ship Type",
    "Operator Domicile",
    "Operator",
    "Group Owner Domicile",
    "Group Owner",
    "Build Location",
    "Build Year",
]

# Segment naming used only when the input is still in subfolders.
SEGMENT_ORDER = ["Bulk", "GC", "Inland", "Misc", "Offshore", "Tankers"]
SEGMENT_OUTPUT_PREFIX = {
    "bulk": "Bulk",
    "gc": "GC",
    "inland": "Inland",
    "misc": "MISC",
    "offshore": "Offshore",
    "tankers": "TANKERS",
}

# The March workbook has Inland.xlsx, not Inland 1.xlsx, because it is one file.
SINGLE_FILE_WITHOUT_NUMBER = {"inland"}


# ============================================================
# HELPERS
# ============================================================

def natural_key(value: str):
    return [int(x) if x.isdigit() else x.lower() for x in re.split(r"(\d+)", value)]


def clean_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [clean_header(c) for c in df.columns]

    # Harmonise common naming differences.
    rename_map = {
        "Source Name": "Source.Name",
        "Sourcename": "Source.Name",
        "Source": "Source.Name",
        "IMO/LR/HS No.": "IMO/LR/IHS No.",
        "IMO LR IHS No.": "IMO/LR/IHS No.",
        "IMO No.": "IMO/LR/IHS No.",
        "IMO": "IMO/LR/IHS No.",
        "Built": "Build Year",
        "Build Date": "Build Year",
    }
    df.rename(columns={c: rename_map.get(c, c) for c in df.columns}, inplace=True)
    return df


def remove_patrol_vessels(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Remove patrol vessels because naval/warship tonnage is out of scope."""
    if "Ship Type" not in df.columns:
        return df, 0

    ship_type = (
        df["Ship Type"]
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .str.lower()
    )

    mask_patrol = ship_type.str.contains(r"\bpatrol\s+vessels?\b", regex=True, na=False)
    removed = int(mask_patrol.sum())
    return df.loc[~mask_patrol].copy(), removed


def read_excel_file(path: Path) -> pd.DataFrame:
    """Read one Seaweb/IHS file using row 2 as header, matching the Power Query logic."""
    xls = pd.ExcelFile(path)
    sheet_name = PREFERRED_INPUT_SHEET if PREFERRED_INPUT_SHEET in xls.sheet_names else xls.sheet_names[0]

    # header=1 means row 2 becomes headers and row 1 is skipped.
    df = pd.read_excel(path, sheet_name=sheet_name, header=1, dtype=object)
    df = normalize_columns(df)

    # Remove empty export rows if present.
    df.dropna(how="all", inplace=True)
    return df


def detect_input_files(input_folder: Path) -> list[tuple[Path, str]]:
    """
    Return list of (file_path, source_name_for_output).

    If direct .xlsx files exist in INPUT_FOLDER, preserve their filenames.
    If files are in segment subfolders, generate names like Bulk 1.xlsx, GC 1.xlsx,
    MISC 1.xlsx, TANKERS 1.xlsx, etc.
    """
    if not input_folder.exists():
        sys.exit(f"INPUT_FOLDER does not exist: {input_folder}")

    direct_files = sorted(
        [f for f in input_folder.glob("*.xlsx") if not f.name.startswith("~$")],
        key=lambda p: natural_key(p.name),
    )

    # Flat folder case: the folder already looks like the target folder.
    if direct_files:
        return [(f, f.name) for f in direct_files if f.resolve() != OUT_PATH.resolve()]

    # Raw segmented folder case.
    files_by_segment: dict[str, list[Path]] = defaultdict(list)

    for folder in sorted([p for p in input_folder.iterdir() if p.is_dir()], key=lambda p: natural_key(p.name)):
        segment_key = folder.name.strip().lower()
        for f in sorted(folder.glob("*.xlsx"), key=lambda p: natural_key(p.name)):
            if not f.name.startswith("~$"):
                files_by_segment[segment_key].append(f)

    if not files_by_segment:
        sys.exit(f"No .xlsx files found in: {input_folder}")

    ordered_segment_keys = []
    for seg in SEGMENT_ORDER:
        key = seg.lower()
        if key in files_by_segment:
            ordered_segment_keys.append(key)
    for key in sorted(files_by_segment.keys(), key=natural_key):
        if key not in ordered_segment_keys:
            ordered_segment_keys.append(key)

    result: list[tuple[Path, str]] = []
    for segment_key in ordered_segment_keys:
        files = files_by_segment[segment_key]
        prefix = SEGMENT_OUTPUT_PREFIX.get(segment_key, segment_key.title())

        for i, file_path in enumerate(files, start=1):
            if len(files) == 1 and segment_key in SINGLE_FILE_WITHOUT_NUMBER:
                source_name = f"{prefix}.xlsx"
            else:
                source_name = f"{prefix} {i}.xlsx"
            result.append((file_path, source_name))

    return result


def replace_segment_folders_with_flat_files(input_folder: Path, file_map: list[tuple[Path, str]]) -> bool:
    """
    Replace segment subfolders with renamed files directly in input_folder.

    Example:
    input_folder/Bulk/Ship Results ...xlsx -> input_folder/Bulk 1.xlsx
    """
    def remove_readonly_and_retry(function, path, exc_info):
        os.chmod(path, stat.S_IWRITE)
        function(path)

    def remove_segment_folders(segment_folders: set[Path]) -> bool:
        removed_any = False
        for segment_folder in sorted(segment_folders, key=lambda p: len(p.parts), reverse=True):
            segment_resolved = segment_folder.resolve()
            if segment_resolved == input_root or input_root not in segment_resolved.parents:
                raise RuntimeError(f"Refusing to remove folder outside INPUT_FOLDER: {segment_folder}")
            try:
                shutil.rmtree(segment_folder, onexc=remove_readonly_and_retry)
                removed_any = True
            except PermissionError as exc:
                print(f"Could not delete segment folder, probably because it is open or locked: {segment_folder}")
                print(f"Windows error: {exc}")
        return removed_any

    direct_files = [f for f in input_folder.glob("*.xlsx") if not f.name.startswith("~$")]
    existing_segment_folders = {
        p
        for p in input_folder.iterdir()
        if p.is_dir() and p.name.strip().lower() in {segment.lower() for segment in SEGMENT_ORDER}
    }

    input_root = input_folder.resolve()
    if direct_files:
        return remove_segment_folders(existing_segment_folders)

    temp_folder = input_folder / "__flat_replacement_tmp__"
    segment_folders: set[Path] = set()

    if temp_folder.exists():
        shutil.rmtree(temp_folder, onexc=remove_readonly_and_retry)
    temp_folder.mkdir(parents=True)

    for source_path, source_name in file_map:
        source_resolved = source_path.resolve()
        if input_root not in source_resolved.parents:
            raise RuntimeError(f"Refusing to copy file outside INPUT_FOLDER: {source_path}")

        if source_path.parent.resolve() != input_root:
            segment_folders.add(source_path.parent)

        shutil.copy2(source_path, temp_folder / source_name)

    for staged_file in temp_folder.glob("*.xlsx"):
        target_path = input_folder / staged_file.name
        if target_path.exists():
            target_path.unlink()
        shutil.move(str(staged_file), str(target_path))

    shutil.rmtree(temp_folder, onexc=remove_readonly_and_retry)

    remove_segment_folders(segment_folders)

    return bool(segment_folders)


def coerce_final_types(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Keep IMO as text. Excel otherwise tends to treat it as a number.
    if "IMO/LR/IHS No." in df.columns:
        df["IMO/LR/IHS No."] = df["IMO/LR/IHS No."].astype(str).str.strip()
        df = df[df["IMO/LR/IHS No."].ne("") & df["IMO/LR/IHS No."].ne("nan")].copy()

    for col in ["GT", "Deadweight", "TEU"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            # Keep blank cells blank, but write actual integers where available.
            df[col] = df[col].where(df[col].isna(), df[col].round(0).astype("Int64"))

    if "Build Year" in df.columns:
        df["Build Year"] = df["Build Year"].astype(str).replace({"nan": ""})

    return df


def format_output_worksheet(ws) -> None:
    # Match the simple one-sheet workbook style: header row + data.
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    widths = {
        "A": 13,
        "B": 13,
        "C": 49,
        "D": 30,
        "E": 13,
        "F": 13,
        "G": 13,
        "H": 13,
        "I": 25,
        "J": 13,
        "K": 13,
        "L": 13,
        "M": 13,
        "N": 13,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # Treat IMO and Build Year as text columns.
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = "@"
        ws.cell(row=row, column=14).number_format = "@"

    # No filter, no freeze panes, no charts, no pivot sheets. One clean data sheet.
    ws.auto_filter.ref = None
    ws.freeze_panes = None


# ============================================================
# MAIN FLOW
# ============================================================

def main() -> None:
    file_map = detect_input_files(INPUT_FOLDER)

    flattened_input = replace_segment_folders_with_flat_files(INPUT_FOLDER, file_map)
    if flattened_input:
        file_map = detect_input_files(INPUT_FOLDER)

    frames = []
    failed_files = []

    for file_path, source_name in file_map:
        try:
            df = read_excel_file(file_path)
            df.drop(columns=["Source.Name", "Source Name"], errors="ignore", inplace=True)
            df.insert(0, "Source.Name", source_name)
            frames.append(df)
        except Exception as exc:
            failed_files.append((str(file_path), str(exc)))

    if not frames:
        sys.exit("No readable input files. Stopping.")

    combined = pd.concat(frames, ignore_index=True)
    combined = normalize_columns(combined)

    rows_before_scope_filter = len(combined)
    combined, patrol_removed = remove_patrol_vessels(combined)
    rows_after_scope_filter = len(combined)

    # Keep only final columns and create missing columns as blanks.
    for col in OUTPUT_COLUMNS:
        if col not in combined.columns:
            combined[col] = ""
    combined = combined[OUTPUT_COLUMNS]

    combined = coerce_final_types(combined)

    rows_before_dedup = len(combined)
    if "IMO/LR/IHS No." in combined.columns:
        combined = combined.drop_duplicates(subset=["IMO/LR/IHS No."], keep="first").copy()
    rows_after_dedup = len(combined)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    if OUT_PATH.exists():
        OUT_PATH.unlink()

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME, index=False)
        format_output_worksheet(writer.sheets[OUTPUT_SHEET_NAME])

    print("Done")
    print(f"Input files read: {len(file_map)}")
    print(f"Rows before Patrol Vessel filter: {rows_before_scope_filter}")
    print(f"Patrol Vessel rows removed: {patrol_removed}")
    print(f"Rows after Patrol Vessel filter: {rows_after_scope_filter}")
    print(f"Rows before IMO deduplication: {rows_before_dedup}")
    print(f"IMO duplicates removed: {rows_before_dedup - rows_after_dedup}")
    print(f"Final rows: {rows_after_dedup}")
    print(f"Output workbook: {OUT_PATH}")
    if flattened_input:
        print(f"Input folder flattened: {INPUT_FOLDER}")

    if failed_files:
        print("Files skipped due to errors:")
        for file_name, error in failed_files:
            print(f"- {file_name}: {error}")


if __name__ == "__main__":
    main()
