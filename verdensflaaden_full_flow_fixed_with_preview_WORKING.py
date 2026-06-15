from __future__ import annotations

import re
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION - EDIT THESE PATHS
# ============================================================

INPUT_FOLDER = Path(r"C:\Users\mrt\OneDrive - Danske Rederier\Desktop\Verdensflåden\2025.12.05")

# Final workbook. Keep this outside INPUT_FOLDER if possible.
OUT_PATH = Path(r"W:\Data\Verdensflaaden\2026\Verdensflaaden.2026.03.14.xlsx")

# Separate workbook for checking the figures before writing to Faktaark.xlsx.
CREATE_FACTAARK_PREVIEW = True
FACTAARK_PREVIEW_PATH = OUT_PATH.with_name(f"{OUT_PATH.stem}_faktaark_preview.xlsx")
TOP_N = 20

PREFERRED_INPUT_SHEET = "Sheet"
OUTPUT_SHEET_NAME = "Sheet1"
MAX_READ_WORKERS = 4

# The final workbook must have exactly these columns and this order.
OUTPUT_COLUMNS = [
    "Source.Name",
    "IMO/LR/IHS No.",
    "Name of Ship",
    "Flag",
    "GT",
    "Deadweight",
    "TEU",
    "Ship Type",
    "Operator Domicile",
    "Operator",
    "Group Owner Domicile",
    "Group Owner",
    "Build Location",
    "Build Year",
]

# Segment naming used only when the input is still in subfolders.
SEGMENT_ORDER = ["Bulk", "GC", "Inland", "Misc", "Offshore", "Tankers"]
SEGMENT_OUTPUT_PREFIX = {
    "bulk": "Bulk",
    "gc": "GC",
    "inland": "Inland",
    "misc": "MISC",
    "offshore": "Offshore",
    "tankers": "TANKERS",
}

# The March workbook has Inland.xlsx, not Inland 1.xlsx, because it is one file.
SINGLE_FILE_WITHOUT_NUMBER = {"inland"}


# ============================================================
# HELPERS
# ============================================================

def natural_key(value: str):
    """Sort strings naturally, e.g. Bulk 2 before Bulk 10."""
    return tuple(int(x) if x.isdigit() else x.lower() for x in re.split(r"(\d+)", str(value)))


def clean_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [clean_header(c) for c in df.columns]

    rename_map = {
        "Source Name": "Source.Name",
        "Sourcename": "Source.Name",
        "Source": "Source.Name",
        "IMO/LR/HS No.": "IMO/LR/IHS No.",
        "IMO LR IHS No.": "IMO/LR/IHS No.",
        "IMO No.": "IMO/LR/IHS No.",
        "IMO": "IMO/LR/IHS No.",
        "Built": "Build Year",
        "Build Date": "Build Year",
    }
    df.rename(columns={c: rename_map.get(c, c) for c in df.columns}, inplace=True)
    return df


def remove_patrol_vessels(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Remove patrol vessels because naval/warship tonnage is out of scope."""
    if "Ship Type" not in df.columns:
        return df, 0

    ship_type = (
        df["Ship Type"]
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .str.lower()
    )

    mask_patrol = ship_type.str.contains(r"\bpatrol\s+vessels?\b", regex=True, na=False)
    removed = int(mask_patrol.sum())
    return df.loc[~mask_patrol].copy(), removed


def read_excel_file(path: Path) -> pd.DataFrame:
    """Read one Seaweb/IHS file using row 2 as header, matching the Power Query logic."""
    with pd.ExcelFile(path, engine="openpyxl") as xls:
        sheet_name = PREFERRED_INPUT_SHEET if PREFERRED_INPUT_SHEET in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet_name, header=1, dtype=object)

    df = normalize_columns(df)
    df.dropna(how="all", inplace=True)
    return df


def read_source_frame(file_path: Path, source_name: str) -> pd.DataFrame:
    df = read_excel_file(file_path)
    df.drop(columns=["Source.Name", "Source Name"], errors="ignore", inplace=True)
    df.insert(0, "Source.Name", source_name)
    return df


def detect_input_files(input_folder: Path) -> list[tuple[Path, str]]:
    """
    Return list of (file_path, source_name_for_output).

    If files are in segment subfolders, generate names like Bulk 1.xlsx, GC 1.xlsx,
    MISC 1.xlsx, TANKERS 1.xlsx, etc.
    If no segment subfolders exist, direct .xlsx files in INPUT_FOLDER are used.
    """
    if not input_folder.exists():
        sys.exit(f"INPUT_FOLDER does not exist: {input_folder}")

    files_by_segment: dict[str, list[Path]] = defaultdict(list)

    for folder in sorted([p for p in input_folder.iterdir() if p.is_dir()], key=lambda p: natural_key(p.name)):
        segment_key = folder.name.strip().lower()
        for f in sorted(folder.glob("*.xlsx"), key=lambda p: natural_key(p.name)):
            if not f.name.startswith("~$"):
                files_by_segment[segment_key].append(f)

    if not files_by_segment:
        direct_files = sorted(
            [f for f in input_folder.glob("*.xlsx") if not f.name.startswith("~$")],
            key=lambda p: natural_key(p.name),
        )
        if direct_files:
            return [(f, f.name) for f in direct_files if f.resolve() != OUT_PATH.resolve()]
        sys.exit(f"No .xlsx files found in: {input_folder}")

    ordered_segment_keys: list[str] = []
    for seg in SEGMENT_ORDER:
        key = seg.lower()
        if key in files_by_segment:
            ordered_segment_keys.append(key)

    for key in sorted(files_by_segment.keys(), key=natural_key):
        if key not in ordered_segment_keys:
            ordered_segment_keys.append(key)

    result: list[tuple[Path, str]] = []
    for segment_key in ordered_segment_keys:
        files = files_by_segment[segment_key]
        prefix = SEGMENT_OUTPUT_PREFIX.get(segment_key, segment_key.title())

        for i, file_path in enumerate(files, start=1):
            if len(files) == 1 and segment_key in SINGLE_FILE_WITHOUT_NUMBER:
                source_name = f"{prefix}.xlsx"
            else:
                source_name = f"{prefix} {i}.xlsx"
            result.append((file_path, source_name))

    return result


def coerce_final_types(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "IMO/LR/IHS No." in df.columns:
        df["IMO/LR/IHS No."] = df["IMO/LR/IHS No."].astype(str).str.strip()
        df = df[df["IMO/LR/IHS No."].ne("") & df["IMO/LR/IHS No."].ne("nan")].copy()

    for col in ["GT", "Deadweight", "TEU"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            df[col] = df[col].where(df[col].isna(), df[col].round(0).astype("Int64"))

    if "Build Year" in df.columns:
        df["Build Year"] = df["Build Year"].astype(str).replace({"nan": ""})

    return df


def format_output_worksheet(ws) -> None:
    """Format the final world fleet sheet."""
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    widths = {
        "A": 13,
        "B": 13,
        "C": 49,
        "D": 30,
        "E": 13,
        "F": 13,
        "G": 13,
        "H": 13,
        "I": 25,
        "J": 13,
        "K": 13,
        "L": 13,
        "M": 13,
        "N": 13,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = "@"
        ws.cell(row=row, column=14).number_format = "@"

    ws.auto_filter.ref = None
    ws.freeze_panes = None


# ============================================================
# FACTAARK PREVIEW
# ============================================================

def normalize_country_name(value: object) -> str:
    """Normalise country names used in flag/operator tables."""
    if value is None or pd.isna(value):
        return ""

    text = str(value).replace("\xa0", " ").strip()
    if not text or text.lower() in {"nan", "none"}:
        return ""

    text = re.sub(r"\s*\((dis|nis|mar|fis|fas|das)\)\s*$", "", text, flags=re.IGNORECASE)
    return text.strip()


def derive_segment_from_source(source_name: object) -> str:
    """Derive broad vessel segment from Source.Name."""
    if source_name is None or pd.isna(source_name):
        return "Unknown"

    name = str(source_name).replace("\xa0", " ").strip().lower()

    if name.startswith("bulk"):
        return "Bulk"
    if name.startswith("gc"):
        return "GC"
    if name.startswith("inland"):
        return "Inland"
    if name.startswith("misc"):
        return "Misc"
    if name.startswith("offshore"):
        return "Offshore"
    if name.startswith("tankers"):
        return "Tankers"

    return "Unknown"


def prepare_preview_data(df: pd.DataFrame) -> pd.DataFrame:
    """Add analysis columns used only in the separate preview workbook."""
    out = df.copy()

    for col in ["GT", "Deadweight", "TEU"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)
        else:
            out[col] = 0

    if "Source.Name" in out.columns:
        out["Segment"] = out["Source.Name"].apply(derive_segment_from_source)
    else:
        out["Segment"] = "Unknown"

    if "Flag" in out.columns:
        out["Flag_Normalized"] = out["Flag"].apply(normalize_country_name)
    else:
        out["Flag_Normalized"] = ""

    if "Operator Domicile" in out.columns:
        out["Operator_Domicile_Normalized"] = out["Operator Domicile"].apply(normalize_country_name)
    else:
        out["Operator_Domicile_Normalized"] = ""

    out["Is_DK_Operated"] = out["Operator_Domicile_Normalized"].str.casefold().eq("denmark")
    out["Is_DK_Flag"] = out["Flag_Normalized"].str.casefold().eq("denmark")

    return out


def make_rank_table(df: pd.DataFrame, group_col: str, display_col: str) -> pd.DataFrame:
    """Top countries by GT for operated fleet and flag fleet."""
    columns = ["Rank", display_col, "Ships", "GT", "Mio. GT", "Deadweight", "TEU", "Label"]
    if group_col not in df.columns:
        return pd.DataFrame(columns=columns)

    tmp = df.copy()
    tmp[group_col] = tmp[group_col].fillna("").astype(str).str.strip()
    tmp = tmp[~tmp[group_col].str.casefold().isin({"", "unknown", "nan", "none"})]

    if tmp.empty:
        return pd.DataFrame(columns=columns)

    result = (
        tmp.groupby(group_col, dropna=False)
        .agg(
            Ships=("IMO/LR/IHS No.", "count"),
            GT=("GT", "sum"),
            Deadweight=("Deadweight", "sum"),
            TEU=("TEU", "sum"),
        )
        .reset_index()
        .rename(columns={group_col: display_col})
        .sort_values("GT", ascending=False)
        .head(TOP_N)
        .reset_index(drop=True)
    )

    result.insert(0, "Rank", range(1, len(result) + 1))
    result["Mio. GT"] = result["GT"] / 1_000_000
    result["Label"] = result["Rank"].astype(str) + " " + result[display_col].astype(str)

    return result[columns]


def make_breakdown(df: pd.DataFrame, group_col: str, label_col: str) -> pd.DataFrame:
    """Generic breakdown by segment or ship type."""
    columns = [label_col, "Ships", "GT", "Mio. GT", "Deadweight", "TEU", "Andel af GT"]
    if df.empty or group_col not in df.columns:
        return pd.DataFrame(columns=columns)

    tmp = df.copy()
    tmp[group_col] = tmp[group_col].fillna("Unknown").astype(str).str.strip()
    tmp[group_col] = tmp[group_col].replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})

    result = (
        tmp.groupby(group_col, dropna=False)
        .agg(
            Ships=("IMO/LR/IHS No.", "count"),
            GT=("GT", "sum"),
            Deadweight=("Deadweight", "sum"),
            TEU=("TEU", "sum"),
        )
        .reset_index()
        .rename(columns={group_col: label_col})
    )

    if group_col == "Segment":
        order = {name: i for i, name in enumerate(["Bulk", "GC", "Inland", "Misc", "Offshore", "Tankers", "Unknown"])}
        result["_sort"] = result[label_col].map(order).fillna(999)
        result = result.sort_values(["_sort", label_col]).drop(columns="_sort")
    else:
        result = result.sort_values("GT", ascending=False)

    result["Mio. GT"] = result["GT"] / 1_000_000
    total_gt = result["GT"].sum()
    result["Andel af GT"] = result["GT"] / total_gt if total_gt else 0

    total_row = pd.DataFrame([{
        label_col: "Total",
        "Ships": result["Ships"].sum(),
        "GT": result["GT"].sum(),
        "Mio. GT": result["GT"].sum() / 1_000_000,
        "Deadweight": result["Deadweight"].sum(),
        "TEU": result["TEU"].sum(),
        "Andel af GT": 1 if total_gt else 0,
    }])

    return pd.concat([result[columns], total_row], ignore_index=True)


def make_source_control(df: pd.DataFrame) -> pd.DataFrame:
    """Control table: files read and how they map to broad segments."""
    columns = ["Source.Name", "Segment", "Ships", "GT", "Mio. GT", "Deadweight", "TEU"]
    if "Source.Name" not in df.columns:
        return pd.DataFrame(columns=columns)

    result = (
        df.groupby(["Source.Name", "Segment"], dropna=False)
        .agg(
            Ships=("IMO/LR/IHS No.", "count"),
            GT=("GT", "sum"),
            Deadweight=("Deadweight", "sum"),
            TEU=("TEU", "sum"),
        )
        .reset_index()
        .sort_values(["Segment", "Source.Name"], key=lambda s: s.map(lambda x: natural_key(str(x))))
    )
    result["Mio. GT"] = result["GT"] / 1_000_000
    return result[columns]


def make_top20_gt_table(df: pd.DataFrame, group_col: str, label_col: str) -> pd.DataFrame:
    """Small two-column table for a simple Top 20 preview sheet."""
    if group_col not in df.columns or "GT" not in df.columns:
        return pd.DataFrame(columns=[label_col, "GT mio"])

    work = df[[group_col, "GT"]].copy()
    work[group_col] = work[group_col].astype(str).str.replace("\xa0", " ", regex=False).str.strip()
    work = work[work[group_col].ne("") & work[group_col].str.lower().ne("nan")].copy()
    work["GT"] = pd.to_numeric(work["GT"], errors="coerce").fillna(0)

    top20 = (
        work.groupby(group_col, as_index=False)["GT"]
        .sum()
        .sort_values(["GT", group_col], ascending=[False, True])
        .head(TOP_N)
        .copy()
    )
    top20["GT mio"] = (top20["GT"] / 1_000_000).round(0).astype("Int64")
    top20[label_col] = [f"{rank} {name}" for rank, name in enumerate(top20[group_col], start=1)]
    return top20[[label_col, "GT mio"]]


def write_top20_block(ws, start_row: int, start_col: int, title: str, table: pd.DataFrame) -> None:
    """Write a formatted Top 20 block."""
    thin_gray = Side(style="thin", color="D9D9D9")
    medium_black = Side(style="medium", color="000000")

    label_col = table.columns[0]
    value_col = table.columns[1]
    end_row = start_row + 21
    end_col = start_col + 1

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(start_row, start_col, title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(start_row + 1, start_col, label_col)
    ws.cell(start_row + 1, start_col + 1, value_col)
    ws.cell(start_row + 1, start_col).font = Font(bold=True)
    ws.cell(start_row + 1, start_col + 1).font = Font(bold=True)

    for index, row in table.reset_index(drop=True).iterrows():
        excel_row = start_row + 2 + index
        ws.cell(excel_row, start_col, row[label_col])
        value = row[value_col]
        ws.cell(excel_row, start_col + 1, None if pd.isna(value) else int(value))

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            top = medium_black if row == start_row else thin_gray
            bottom = medium_black if row in {start_row, end_row} else thin_gray
            left = medium_black if col == start_col else thin_gray
            right = medium_black if col == end_col else thin_gray
            if row == start_row:
                bottom = medium_black

            cell = ws.cell(row, col)
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
            cell.alignment = Alignment(
                horizontal="right" if col == start_col + 1 and row >= start_row + 2 else "left",
                vertical="center",
            )

    ws.cell(start_row, start_col).alignment = Alignment(horizontal="center", vertical="center")


def add_top20_sheet_to_workbook(path: Path, df: pd.DataFrame) -> None:
    """Add or replace a formatted Top 20 sheet in a workbook."""
    operator_table = make_top20_gt_table(df, "Operator Domicile", "Operatørland")
    flag_table = make_top20_gt_table(df, "Flag", "Flag")

    wb = load_workbook(path)

    if "Top 20" in wb.sheetnames:
        del wb["Top 20"]

    ws = wb.create_sheet("Top 20", 0)
    wb.active = 0
    ws.sheet_view.showGridLines = True
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 12

    write_top20_block(ws, 1, 1, "20 største operatører", operator_table)
    write_top20_block(ws, 1, 5, "20 største flagstater", flag_table)

    wb.save(path)


def write_preview_section(writer, sheet_name: str, title: str, df: pd.DataFrame, startrow: int, startcol: int = 0) -> int:
    """Write one titled dataframe section and return the next available row."""
    df.to_excel(writer, sheet_name=sheet_name, startrow=startrow + 1, startcol=startcol, index=False)
    ws = writer.sheets[sheet_name]

    title_cell = ws.cell(row=startrow + 1, column=startcol + 1)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=12)

    header_row = startrow + 2
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for col_idx in range(startcol + 1, startcol + len(df.columns) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    return startrow + len(df) + 4


def format_preview_worksheet(ws) -> None:
    """Basic formatting for preview workbook."""
    ws.freeze_panes = "A2"

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        width = 14

        for row_idx in range(1, min(ws.max_row, 60) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                width = max(width, min(len(str(value)) + 2, 35))

        ws.column_dimensions[col_letter].width = width

    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, float):
                header = str(ws.cell(row=2 if row_idx > 1 else 1, column=col_idx).value or "")
                if "Andel" in header:
                    cell.number_format = "0.0%"
                elif "Mio" in header:
                    cell.number_format = "0.0"
                else:
                    cell.number_format = "#,##0"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"

    for row_idx in range(1, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        if any(value == "Total" for value in row_values):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)


def create_faktaark_preview(
    df: pd.DataFrame,
    preview_path: Path,
    input_files_read: int,
    rows_before_scope_filter: int,
    patrol_removed: int,
    rows_after_scope_filter: int,
    rows_before_dedup: int,
    rows_after_dedup: int,
    failed_files: list[tuple[str, str]],
) -> None:
    """Create separate Excel workbook showing the numbers that can later feed Faktaark.xlsx."""
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    if preview_path.exists():
        preview_path.unlink()

    preview_df = prepare_preview_data(df)

    top_operator = make_rank_table(preview_df, "Operator_Domicile_Normalized", "Operator Domicile")
    top_flag = make_rank_table(preview_df, "Flag_Normalized", "Flag")

    dk_operated = preview_df[preview_df["Is_DK_Operated"]].copy()
    dk_flag = preview_df[preview_df["Is_DK_Flag"]].copy()

    dk_operated_segments = make_breakdown(dk_operated, "Segment", "Segment")
    dk_flag_segments = make_breakdown(dk_flag, "Segment", "Segment")

    dk_operated_shiptypes = make_breakdown(dk_operated, "Ship Type", "Ship Type")
    dk_flag_shiptypes = make_breakdown(dk_flag, "Ship Type", "Ship Type")

    source_control = make_source_control(preview_df)

    control = pd.DataFrame(
        [
            ["Kørselsdato", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Inputmappe", str(INPUT_FOLDER)],
            ["Verdensflåde-output", str(OUT_PATH)],
            ["Preview-output", str(preview_path)],
            ["Inputfiler læst", input_files_read],
            ["Rækker før Patrol Vessel-filter", rows_before_scope_filter],
            ["Patrol Vessel-rækker fjernet", patrol_removed],
            ["Rækker efter Patrol Vessel-filter", rows_after_scope_filter],
            ["Rækker før IMO-dubletfjernelse", rows_before_dedup],
            ["IMO-dubletter fjernet", rows_before_dedup - rows_after_dedup],
            ["Endelige rækker", rows_after_dedup],
            ["Fejlede filer", len(failed_files)],
        ],
        columns=["Kontrolpunkt", "Værdi"],
    )

    failed = pd.DataFrame(failed_files, columns=["Fil", "Fejl"]) if failed_files else pd.DataFrame(columns=["Fil", "Fejl"])

    with pd.ExcelWriter(preview_path, engine="openpyxl") as writer:
        write_preview_section(writer, "Opereret_flag", "Top 20 operatørlande efter GT", top_operator, startrow=0, startcol=0)
        write_preview_section(writer, "Opereret_flag", "Top 20 flagstater efter GT", top_flag, startrow=0, startcol=10)

        next_row = write_preview_section(writer, "DK_segmenter", "Dansk opereret flåde fordelt på segment", dk_operated_segments, startrow=0, startcol=0)
        write_preview_section(writer, "DK_segmenter", "Dansk flagflåde fordelt på segment", dk_flag_segments, startrow=next_row, startcol=0)

        next_row = write_preview_section(writer, "DK_fartojstyper", "Dansk opereret flåde fordelt på Ship Type", dk_operated_shiptypes, startrow=0, startcol=0)
        write_preview_section(writer, "DK_fartojstyper", "Dansk flagflåde fordelt på Ship Type", dk_flag_shiptypes, startrow=next_row, startcol=0)

        source_control.to_excel(writer, sheet_name="Kildekontrol", index=False)
        control.to_excel(writer, sheet_name="Kontrol", index=False)
        failed.to_excel(writer, sheet_name="Fejlede_filer", index=False)

        for ws in writer.book.worksheets:
            format_preview_worksheet(ws)

    add_top20_sheet_to_workbook(preview_path, df)


# ============================================================
# MAIN FLOW
# ============================================================

def read_all_input_frames(file_map: list[tuple[Path, str]]) -> tuple[list[pd.DataFrame], list[tuple[str, str]]]:
    frames: list[pd.DataFrame] = []
    failed_files: list[tuple[str, str]] = []

    worker_count = max(1, min(MAX_READ_WORKERS, len(file_map)))
    print(f"Reading input files: {len(file_map)} file(s), {worker_count} worker(s)")

    if worker_count == 1:
        for file_path, source_name in file_map:
            try:
                frames.append(read_source_frame(file_path, source_name))
            except Exception as exc:
                failed_files.append((str(file_path), str(exc)))
        return frames, failed_files

    ordered_frames: list[pd.DataFrame | None] = [None] * len(file_map)
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(read_source_frame, file_path, source_name): (index, file_path)
            for index, (file_path, source_name) in enumerate(file_map)
        }

        for future in as_completed(futures):
            index, file_path = futures[future]
            try:
                ordered_frames[index] = future.result()
            except Exception as exc:
                failed_files.append((str(file_path), str(exc)))

    frames = [frame for frame in ordered_frames if frame is not None]
    return frames, failed_files


def build_world_fleet(file_map: list[tuple[Path, str]]) -> tuple[pd.DataFrame, dict[str, int], list[tuple[str, str]]]:
    frames, failed_files = read_all_input_frames(file_map)

    if not frames:
        sys.exit("No readable input files. Stopping.")

    combined = pd.concat(frames, ignore_index=True)
    combined = normalize_columns(combined)

    rows_before_scope_filter = len(combined)
    combined, patrol_removed = remove_patrol_vessels(combined)
    rows_after_scope_filter = len(combined)

    for col in OUTPUT_COLUMNS:
        if col not in combined.columns:
            combined[col] = ""

    combined = combined[OUTPUT_COLUMNS]
    combined = coerce_final_types(combined)

    rows_before_dedup = len(combined)
    if "IMO/LR/IHS No." in combined.columns:
        combined = combined.drop_duplicates(subset=["IMO/LR/IHS No."], keep="first").copy()
    rows_after_dedup = len(combined)

    stats = {
        "rows_before_scope_filter": rows_before_scope_filter,
        "patrol_removed": patrol_removed,
        "rows_after_scope_filter": rows_after_scope_filter,
        "rows_before_dedup": rows_before_dedup,
        "rows_after_dedup": rows_after_dedup,
    }

    return combined, stats, failed_files


def write_world_fleet_output(combined: pd.DataFrame) -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    if OUT_PATH.exists():
        OUT_PATH.unlink()

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME, index=False)
        format_output_worksheet(writer.sheets[OUTPUT_SHEET_NAME])


def main() -> None:
    file_map = detect_input_files(INPUT_FOLDER)
    combined, stats, failed_files = build_world_fleet(file_map)

    write_world_fleet_output(combined)

    if CREATE_FACTAARK_PREVIEW:
        create_faktaark_preview(
            combined,
            FACTAARK_PREVIEW_PATH,
            input_files_read=len(file_map),
            rows_before_scope_filter=stats["rows_before_scope_filter"],
            patrol_removed=stats["patrol_removed"],
            rows_after_scope_filter=stats["rows_after_scope_filter"],
            rows_before_dedup=stats["rows_before_dedup"],
            rows_after_dedup=stats["rows_after_dedup"],
            failed_files=failed_files,
        )

    print("Done")
    print(f"Input files read: {len(file_map)}")
    print(f"Rows before Patrol Vessel filter: {stats['rows_before_scope_filter']}")
    print(f"Patrol Vessel rows removed: {stats['patrol_removed']}")
    print(f"Rows after Patrol Vessel filter: {stats['rows_after_scope_filter']}")
    print(f"Rows before IMO deduplication: {stats['rows_before_dedup']}")
    print(f"IMO duplicates removed: {stats['rows_before_dedup'] - stats['rows_after_dedup']}")
    print(f"Final rows: {stats['rows_after_dedup']}")
    print(f"Output workbook: {OUT_PATH}")

    if CREATE_FACTAARK_PREVIEW:
        print(f"Faktaark preview workbook: {FACTAARK_PREVIEW_PATH}")

    if failed_files:
        print("Files skipped due to errors:")
        for file_name, error in failed_files:
            print(f"- {file_name}: {error}")


if __name__ == "__main__":
    main()
