# -*- coding: utf-8 -*-
from __future__ import annotations

import re
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import pandas as pd
from pathlib import Path as _Path

def _resolve_faktaark_try_path(path):
    raw_path = _Path(path)
    candidates = [
        raw_path,
        _Path(__file__).resolve().parent / raw_path.name,
        _Path.cwd() / raw_path.name,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[1]
from copy import copy as _copy

def _top20_ranges_intersect(range_a, range_b):
    return not (
        range_a.max_row < range_b.min_row
        or range_a.min_row > range_b.max_row
        or range_a.max_col < range_b.min_col
        or range_a.min_col > range_b.max_col
    )

def _copy_top20_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell._style = _copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.font:
        target_cell.font = _copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = _copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = _copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = _copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = _copy(source_cell.protection)
    if source_cell.hyperlink:
        target_cell._hyperlink = _copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = _copy(source_cell.comment)

def _find_top20_table_title(sheet, title):
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value).strip() == title:
                return cell
    return None

def _top20_source_range(sheet, title_cell):
    from openpyxl.worksheet.cell_range import CellRange

    title_row = title_cell.row
    first_col = title_cell.column
    header_row = title_row + 1
    table_rows = int(globals().get("TOP_TABLE_N", 20)) + 2
    table_width = 2

    title_range = CellRange(
        min_row=title_row,
        max_row=title_row,
        min_col=first_col,
        max_col=first_col,
    )
    for merged_range in sheet.merged_cells.ranges:
        if title_cell.coordinate in merged_range:
            title_range = merged_range
            table_width = max(table_width, merged_range.max_col - merged_range.min_col + 1)
            first_col = merged_range.min_col
            break

    header_width = 0
    for col in range(first_col, sheet.max_column + 1):
        if sheet.cell(header_row, col).value is None and header_width:
            break
        if sheet.cell(header_row, col).value is not None:
            header_width = col - first_col + 1
    table_width = max(table_width, header_width, 2)

    return CellRange(
        min_row=title_row,
        max_row=title_row + table_rows - 1,
        min_col=first_col,
        max_col=first_col + table_width - 1,
    )

def _copy_top20_table(source_ws, target_ws, title, dest_row, dest_col):
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import CellRange

    title_cell = _find_top20_table_title(source_ws, title)
    if title_cell is None:
        return

    source_range = _top20_source_range(source_ws, title_cell)
    dest_range = CellRange(
        min_row=dest_row,
        max_row=dest_row + source_range.max_row - source_range.min_row,
        min_col=dest_col,
        max_col=dest_col + source_range.max_col - source_range.min_col,
    )

    for merged_range in list(target_ws.merged_cells.ranges):
        if _top20_ranges_intersect(merged_range, dest_range):
            target_ws.unmerge_cells(str(merged_range))

    for row in target_ws.iter_rows(
        min_row=dest_range.min_row,
        max_row=dest_range.max_row,
        min_col=dest_range.min_col,
        max_col=dest_range.max_col,
    ):
        for cell in row:
            cell.value = None

    for row in source_ws.iter_rows(
        min_row=source_range.min_row,
        max_row=source_range.max_row,
        min_col=source_range.min_col,
        max_col=source_range.max_col,
    ):
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            target_cell = target_ws.cell(
                row=dest_row + source_cell.row - source_range.min_row,
                column=dest_col + source_cell.column - source_range.min_col,
            )
            _copy_top20_cell(source_cell, target_cell)

    for merged_range in source_ws.merged_cells.ranges:
        if _top20_ranges_intersect(merged_range, source_range):
            moved_range = CellRange(
                min_row=dest_row + merged_range.min_row - source_range.min_row,
                max_row=dest_row + merged_range.max_row - source_range.min_row,
                min_col=dest_col + merged_range.min_col - source_range.min_col,
                max_col=dest_col + merged_range.max_col - source_range.min_col,
            )
            target_ws.merge_cells(str(moved_range))

    for source_col in range(source_range.min_col, source_range.max_col + 1):
        source_letter = get_column_letter(source_col)
        target_letter = get_column_letter(dest_col + source_col - source_range.min_col)
        source_width = source_ws.column_dimensions[source_letter].width
        if source_width:
            target_ws.column_dimensions[target_letter].width = source_width

    for source_row in range(source_range.min_row, source_range.max_row + 1):
        source_height = source_ws.row_dimensions[source_row].height
        if source_height:
            target_ws.row_dimensions[dest_row + source_row - source_range.min_row].height = source_height

def _merge_top20_tables_into_top20_sheet(workbook):
    if "Top 20" not in workbook.sheetnames or "Top 20 tabeller" not in workbook.sheetnames:
        return

    chart_ws = workbook["Top 20"]
    table_ws = workbook["Top 20 tabeller"]
    dest_start_row = 25

    _copy_top20_table(table_ws, chart_ws, "20 største operatører", dest_start_row, 2)
    _copy_top20_table(table_ws, chart_ws, "20 største flagstater", dest_start_row, 11)
    workbook.remove(table_ws)

_SEGMENTFORDELING_SECTION_PREFIXES = (
    "Tonnage og fordeling på fartøjs",
    "Opereret flåde og flagflåde (in",
)

def _cell_text(cell):
    return "" if cell.value is None else str(cell.value).strip()

def _find_cell_starting_with(sheet, prefix):
    for row in sheet.iter_rows():
        for cell in row:
            if _cell_text(cell).startswith(prefix):
                return cell
    return None

def _segmentfordeling_section_range(sheet, title_cell):
    from openpyxl.worksheet.cell_range import CellRange

    first_row = title_cell.row
    first_col = title_cell.column
    max_row = first_row
    max_col = first_col

    for row in range(first_row, sheet.max_row + 1):
        has_value = any(sheet.cell(row, col).value is not None for col in range(first_col, sheet.max_column + 1))
        if not has_value and row > first_row:
            break
        if has_value:
            max_row = row
            row_max_col = max(
                col
                for col in range(first_col, sheet.max_column + 1)
                if sheet.cell(row, col).value is not None
            )
            max_col = max(max_col, row_max_col)

    for merged_range in sheet.merged_cells.ranges:
        if title_cell.coordinate in merged_range:
            first_col = min(first_col, merged_range.min_col)
            max_col = max(max_col, merged_range.max_col)
            break

    return CellRange(min_row=first_row, max_row=max_row, min_col=first_col, max_col=max_col)

def _copy_segmentfordeling_section(source_ws, target_ws, prefix):
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import CellRange

    source_title = _find_cell_starting_with(source_ws, prefix)
    if source_title is None:
        return

    source_range = _segmentfordeling_section_range(source_ws, source_title)
    target_title = _find_cell_starting_with(target_ws, prefix)
    dest_row = target_title.row if target_title is not None else source_range.min_row
    dest_col = target_title.column if target_title is not None else source_range.min_col

    dest_range = CellRange(
        min_row=dest_row,
        max_row=dest_row + source_range.max_row - source_range.min_row,
        min_col=dest_col,
        max_col=dest_col + source_range.max_col - source_range.min_col,
    )

    for merged_range in list(target_ws.merged_cells.ranges):
        if _top20_ranges_intersect(merged_range, dest_range):
            target_ws.unmerge_cells(str(merged_range))

    for row in target_ws.iter_rows(
        min_row=dest_range.min_row,
        max_row=dest_range.max_row,
        min_col=dest_range.min_col,
        max_col=dest_range.max_col,
    ):
        for cell in row:
            cell.value = None

    for row in source_ws.iter_rows(
        min_row=source_range.min_row,
        max_row=source_range.max_row,
        min_col=source_range.min_col,
        max_col=source_range.max_col,
    ):
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            target_cell = target_ws.cell(
                row=dest_row + source_cell.row - source_range.min_row,
                column=dest_col + source_cell.column - source_range.min_col,
            )
            _copy_top20_cell(source_cell, target_cell)

    for merged_range in source_ws.merged_cells.ranges:
        if _top20_ranges_intersect(merged_range, source_range):
            moved_range = CellRange(
                min_row=dest_row + merged_range.min_row - source_range.min_row,
                max_row=dest_row + merged_range.max_row - source_range.min_row,
                min_col=dest_col + merged_range.min_col - source_range.min_col,
                max_col=dest_col + merged_range.max_col - source_range.min_col,
            )
            target_ws.merge_cells(str(moved_range))

    for source_col in range(source_range.min_col, source_range.max_col + 1):
        source_letter = get_column_letter(source_col)
        target_letter = get_column_letter(dest_col + source_col - source_range.min_col)
        source_width = source_ws.column_dimensions[source_letter].width
        if source_width:
            target_ws.column_dimensions[target_letter].width = source_width

def _normalize_segmentfordeling_sheet(workbook):
    if "Faktaark try" not in workbook.sheetnames:
        return

    target_ws = workbook["Faktaark try"]
    if "Segmentfordeling" not in workbook.sheetnames:
        return

    source_ws = workbook["Segmentfordeling"]
    for section_prefix in _SEGMENTFORDELING_SECTION_PREFIXES:
        _copy_segmentfordeling_section(source_ws, target_ws, section_prefix)

try:
    from openpyxl.workbook.workbook import Workbook as _OpenpyxlWorkbook

    _original_openpyxl_workbook_save = _OpenpyxlWorkbook.save

    def _save_with_top20_tables_on_chart_sheet(self, filename):
        _merge_top20_tables_into_top20_sheet(self)
        _normalize_segmentfordeling_sheet(self)
        return _original_openpyxl_workbook_save(self, filename)

    if not getattr(_OpenpyxlWorkbook.save, "_top20_tables_on_chart_sheet", False):
        _save_with_top20_tables_on_chart_sheet._top20_tables_on_chart_sheet = True
        _OpenpyxlWorkbook.save = _save_with_top20_tables_on_chart_sheet
except Exception:
    pass
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION - EDIT THESE PATHS
# ============================================================

INPUT_FOLDER = Path(r"C:\Users\mrt\OneDrive - Danske Rederier\Desktop\Verdensflåden\2025.12.05")

# Final workbook. Keep this outside INPUT_FOLDER if possible.
OUT_PATH = Path(r"W:\Data\Verdensflaaden\2026\Verdensflaaden.2026.03.14.xlsx")

# Separate workbook for checking the figures before writing to Faktaark.xlsx.
CREATE_FACTAARK_PREVIEW = True
FACTAARK_PREVIEW_PATH = OUT_PATH.with_name(f"{OUT_PATH.stem}_faktaark_preview.xlsx")

# Optional: update an existing Faktaark try workbook directly.
# First run writes a copy so the original template is not overwritten.
UPDATE_FAKTAARK_TRY = True
FAKTAARK_TRY_PATH = Path(r"C:\Users\mrt\OneDrive - Danske Rederier\Desktop\Verdensflaaden\Faktaark try.xlsx")
FAKTAARK_TRY_OUT_PATH = FAKTAARK_TRY_PATH.with_name(f"{FAKTAARK_TRY_PATH.stem}_opdateret.xlsx")

# Fact-sheet charts use top 15 in the examples.
TOP_N = 15

# The separate table view must still show top 20.
TOP_TABLE_N = 20

PREFERRED_INPUT_SHEET = "Sheet"
OUTPUT_SHEET_NAME = "Sheet1"
MAX_READ_WORKERS = 4

# The final workbook must have exactly these columns and this order.
OUTPUT_COLUMNS = [
    "Source.Name",
    "IMO/LR/IHS No.",
    "Name of Ship",
    "Flag",
    "GT",
    "Deadweight",
    "TEU",
    "Ship Type",
    "Operator Domicile",
    "Operator",
    "Group Owner Domicile",
    "Group Owner",
    "Build Location",
    "Build Year",
]

# Segment naming used only when the input is still in subfolders.
SEGMENT_ORDER = ["Bulk", "GC", "Inland", "Misc", "Offshore", "Tankers"]
SEGMENT_OUTPUT_PREFIX = {
    "bulk": "Bulk",
    "gc": "GC",
    "inland": "Inland",
    "misc": "MISC",
    "offshore": "Offshore",
    "tankers": "TANKERS",
}

# The March workbook has Inland.xlsx, not Inland 1.xlsx, because it is one file.
SINGLE_FILE_WITHOUT_NUMBER = {"inland"}

# Visual style
BLUE = "#155F7A"
ORANGE = "#ED7D31"
GRID_GREY = "#D9D9D9"
TEXT_GREY = "#595959"


# ============================================================
# BASIC HELPERS
# ============================================================

def natural_key(value: str):
    """Sort strings naturally, e.g. Bulk 2 before Bulk 10."""
    return tuple(int(x) if x.isdigit() else x.lower() for x in re.split(r"(\d+)", str(value)))


def clean_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [clean_header(c) for c in df.columns]

    rename_map = {
        "Source Name": "Source.Name",
        "Sourcename": "Source.Name",
        "Source": "Source.Name",
        "IMO/LR/HS No.": "IMO/LR/IHS No.",
        "IMO LR IHS No.": "IMO/LR/IHS No.",
        "IMO No.": "IMO/LR/IHS No.",
        "IMO": "IMO/LR/IHS No.",
        "Built": "Build Year",
        "Build Date": "Build Year",
    }
    df.rename(columns={c: rename_map.get(c, c) for c in df.columns}, inplace=True)
    return df


def normalize_country_name(value: object) -> str:
    """
    Normalise country names.

    Important:
    Denmark, Denmark (DIS), Denmark (DAS), etc. are treated as Denmark.
    Same logic also handles Norway (NIS), etc.
    """
    if value is None or pd.isna(value):
        return ""

    text = str(value).replace("\xa0", " ").strip()
    if not text or text.lower() in {"nan", "none"}:
        return ""

    text = re.sub(r"\s*\((dis|nis|mar|fis|fas|das)\)\s*$", "", text, flags=re.IGNORECASE)
    return text.strip()


def remove_patrol_vessels(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Remove patrol vessels because naval/warship tonnage is out of scope."""
    if "Ship Type" not in df.columns:
        return df, 0

    ship_type = (
        df["Ship Type"]
        .astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .str.lower()
    )

    mask_patrol = ship_type.str.contains(r"\bpatrol\s+vessels?\b", regex=True, na=False)
    removed = int(mask_patrol.sum())
    return df.loc[~mask_patrol].copy(), removed


def read_excel_file(path: Path) -> pd.DataFrame:
    """Read one Seaweb/IHS file using row 2 as header, matching the Power Query logic."""
    with pd.ExcelFile(path, engine="openpyxl") as xls:
        sheet_name = PREFERRED_INPUT_SHEET if PREFERRED_INPUT_SHEET in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet_name, header=1, dtype=object)

    df = normalize_columns(df)
    df.dropna(how="all", inplace=True)
    return df


def read_source_frame(file_path: Path, source_name: str) -> pd.DataFrame:
    df = read_excel_file(file_path)
    df.drop(columns=["Source.Name", "Source Name"], errors="ignore", inplace=True)
    df.insert(0, "Source.Name", source_name)
    return df


def detect_input_files(input_folder: Path) -> list[tuple[Path, str]]:
    """
    Return list of (file_path, source_name_for_output).

    If files are in segment subfolders, generate names like Bulk 1.xlsx, GC 1.xlsx,
    MISC 1.xlsx, TANKERS 1.xlsx, etc.
    If no segment subfolders exist, direct .xlsx files in INPUT_FOLDER are used.
    """
    if not input_folder.exists():
        sys.exit(f"INPUT_FOLDER does not exist: {input_folder}")

    files_by_segment: dict[str, list[Path]] = defaultdict(list)

    for folder in sorted([p for p in input_folder.iterdir() if p.is_dir()], key=lambda p: natural_key(p.name)):
        segment_key = folder.name.strip().lower()
        for f in sorted(folder.glob("*.xlsx"), key=lambda p: natural_key(p.name)):
            if not f.name.startswith("~$"):
                files_by_segment[segment_key].append(f)

    if not files_by_segment:
        direct_files = sorted(
            [f for f in input_folder.glob("*.xlsx") if not f.name.startswith("~$")],
            key=lambda p: natural_key(p.name),
        )
        if direct_files:
            return [(f, f.name) for f in direct_files if f.resolve() != OUT_PATH.resolve()]
        sys.exit(f"No .xlsx files found in: {input_folder}")

    ordered_segment_keys: list[str] = []
    for seg in SEGMENT_ORDER:
        key = seg.lower()
        if key in files_by_segment:
            ordered_segment_keys.append(key)

    for key in sorted(files_by_segment.keys(), key=natural_key):
        if key not in ordered_segment_keys:
            ordered_segment_keys.append(key)

    result: list[tuple[Path, str]] = []
    for segment_key in ordered_segment_keys:
        files = files_by_segment[segment_key]
        prefix = SEGMENT_OUTPUT_PREFIX.get(segment_key, segment_key.title())

        for i, file_path in enumerate(files, start=1):
            if len(files) == 1 and segment_key in SINGLE_FILE_WITHOUT_NUMBER:
                source_name = f"{prefix}.xlsx"
            else:
                source_name = f"{prefix} {i}.xlsx"
            result.append((file_path, source_name))

    return result


def coerce_final_types(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "IMO/LR/IHS No." in df.columns:
        df["IMO/LR/IHS No."] = df["IMO/LR/IHS No."].astype(str).str.strip()
        df = df[df["IMO/LR/IHS No."].ne("") & df["IMO/LR/IHS No."].ne("nan")].copy()

    for col in ["GT", "Deadweight", "TEU"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            df[col] = df[col].where(df[col].isna(), df[col].round(0).astype("Int64"))

    if "Build Year" in df.columns:
        df["Build Year"] = df["Build Year"].astype(str).replace({"nan": ""})

    return df


def format_output_worksheet(ws) -> None:
    """Format the final world fleet sheet."""
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    widths = {
        "A": 13,
        "B": 13,
        "C": 49,
        "D": 30,
        "E": 13,
        "F": 13,
        "G": 13,
        "H": 13,
        "I": 25,
        "J": 13,
        "K": 13,
        "L": 13,
        "M": 13,
        "N": 13,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = "@"
        ws.cell(row=row, column=14).number_format = "@"

    ws.auto_filter.ref = None
    ws.freeze_panes = None


# ============================================================
# PREVIEW DATA PREPARATION
# ============================================================

def derive_segment_from_source(source_name: object) -> str:
    """Derive broad vessel segment from Source.Name."""
    if source_name is None or pd.isna(source_name):
        return "Unknown"

    name = str(source_name).replace("\xa0", " ").strip().lower()

    if name.startswith("bulk"):
        return "Bulk"
    if name.startswith("gc"):
        return "GC"
    if name.startswith("inland"):
        return "Inland"
    if name.startswith("misc"):
        return "Misc"
    if name.startswith("offshore"):
        return "Offshore"
    if name.startswith("tankers"):
        return "Tankers"

    return "Unknown"


def segment_to_fact_sheet_group(segment: object) -> str:
    """Group segments for the lower fact-sheet table and chart."""
    text = str(segment or "").strip().casefold()

    if text == "bulk":
        return "Bulk"
    if text == "gc":
        return "General cargo (incl. Ferries, container and Ro-Ro)"
    if text in {"inland", "misc"}:
        return "Misc (incl. Inland waterways)"
    if text == "offshore":
        return "Offshore"
    if text == "tankers":
        return "Tankers"

    return "Unknown"


def prepare_preview_data(df: pd.DataFrame) -> pd.DataFrame:
    """Add analysis columns used only in the separate preview workbook."""
    out = df.copy()

    for col in ["GT", "Deadweight", "TEU"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)
        else:
            out[col] = 0

    if "Source.Name" in out.columns:
        out["Segment"] = out["Source.Name"].apply(derive_segment_from_source)
    else:
        out["Segment"] = "Unknown"

    out["FactSheetGroup"] = out["Segment"].apply(segment_to_fact_sheet_group)

    if "Flag" in out.columns:
        out["Flag_Normalized"] = out["Flag"].apply(normalize_country_name)
    else:
        out["Flag_Normalized"] = ""

    if "Operator Domicile" in out.columns:
        out["Operator_Domicile_Normalized"] = out["Operator Domicile"].apply(normalize_country_name)
    else:
        out["Operator_Domicile_Normalized"] = ""

    out["Is_DK_Operated"] = out["Operator_Domicile_Normalized"].str.casefold().eq("denmark")
    out["Is_DK_Flag"] = out["Flag_Normalized"].str.casefold().eq("denmark")

    return out


# ============================================================
# TOP 15/TOP 20 TABLES AND CHARTS
# ============================================================

def make_rank_table(df: pd.DataFrame, group_col: str, display_col: str, top_n: int = TOP_N) -> pd.DataFrame:
    """Top countries by GT for operated fleet and flag fleet."""
    columns = ["Rank", display_col, "Ships", "GT", "Mio. GT", "Deadweight", "TEU", "Label"]
    if group_col not in df.columns:
        return pd.DataFrame(columns=columns)

    tmp = df.copy()
    tmp[group_col] = tmp[group_col].fillna("").astype(str).str.strip()
    tmp = tmp[~tmp[group_col].str.casefold().isin({"", "unknown", "nan", "none"})]

    if tmp.empty:
        return pd.DataFrame(columns=columns)

    result = (
        tmp.groupby(group_col, dropna=False)
        .agg(
            Ships=("IMO/LR/IHS No.", "count"),
            GT=("GT", "sum"),
            Deadweight=("Deadweight", "sum"),
            TEU=("TEU", "sum"),
        )
        .reset_index()
        .rename(columns={group_col: display_col})
        .sort_values("GT", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )

    result.insert(0, "Rank", range(1, len(result) + 1))
    result["Mio. GT"] = (result["GT"] / 1_000_000).round(2)
    result["Label"] = result["Rank"].astype(str) + " " + result[display_col].astype(str)

    return result[columns]


def make_top_chart_table(df: pd.DataFrame, group_col: str, label_col: str) -> pd.DataFrame:
    """Create a chart source table: Label + Mio. GT + Country."""
    if group_col not in df.columns or "GT" not in df.columns:
        return pd.DataFrame(columns=[label_col, "Mio. GT", "Country"])

    work = df[[group_col, "GT"]].copy()
    work[group_col] = work[group_col].apply(normalize_country_name)
    work = work[~work[group_col].str.casefold().isin({"", "unknown", "nan", "none"})]
    work["GT"] = pd.to_numeric(work["GT"], errors="coerce").fillna(0)

    top = (
        work.groupby(group_col, as_index=False)["GT"]
        .sum()
        .sort_values(["GT", group_col], ascending=[False, True])
        .head(TOP_N)
        .reset_index(drop=True)
    )

    top["Mio. GT"] = (top["GT"] / 1_000_000).round(2)
    top[label_col] = [f"{rank} {name}" for rank, name in enumerate(top[group_col], start=1)]
    top["Country"] = top[group_col]
    return top[[label_col, "Mio. GT", "Country"]]


def require_matplotlib():
    try:
        import matplotlib.pyplot as plt
        import matplotlib.ticker as mticker
        return plt, mticker
    except ImportError as exc:
        raise RuntimeError(
            "matplotlib mangler. Kør: pip install matplotlib pillow"
        ) from exc


def make_top_horizontal_bar_chart_png(table: pd.DataFrame, title: str, out_png: Path) -> None:
    """Render a top-country horizontal bar chart as PNG."""
    plt, mticker = require_matplotlib()

    if table.empty:
        return

    plot_df = table.copy()
    labels = plot_df.iloc[:, 0].astype(str).tolist()
    values = pd.to_numeric(plot_df["Mio. GT"], errors="coerce").fillna(0).tolist()
    countries = plot_df["Country"].astype(str).tolist()
    colors = [ORANGE if normalize_country_name(c).casefold() == "denmark" else BLUE for c in countries]

    # Keep largest at top by reversing the plotted order.
    labels_r = labels[::-1]
    values_r = values[::-1]
    colors_r = colors[::-1]

    fig_height = max(5.0, 0.36 * len(labels) + 1.4)
    fig, ax = plt.subplots(figsize=(8.8, fig_height), dpi=180)
    bars = ax.barh(labels_r, values_r, color=colors_r, height=0.34)

    ax.set_title(title, fontsize=11, fontweight="bold", color=TEXT_GREY, pad=10)
    ax.set_xlabel("")
    ax.set_ylabel("")
    ax.xaxis.grid(True, color=GRID_GREY, linewidth=0.8)
    ax.yaxis.grid(False)
    ax.set_axisbelow(True)

    max_val = max(values) if values else 0
    ax.set_xlim(0, max_val * 1.18 if max_val else 1)

    for spine in ax.spines.values():
        spine.set_visible(False)

    ax.tick_params(axis="x", colors=TEXT_GREY, labelsize=8)
    ax.tick_params(axis="y", colors=TEXT_GREY, labelsize=8, length=0)
    ax.xaxis.set_major_formatter(mticker.StrMethodFormatter("{x:.0f}"))

    for bar, value in zip(bars, values_r):
        ax.text(
            bar.get_width() + max_val * 0.015,
            bar.get_y() + bar.get_height() / 2,
            f"{value:.2f}".replace(".", ","),
            va="center",
            ha="left",
            fontsize=8,
            color=TEXT_GREY,
        )

    fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, bbox_inches="tight", facecolor="white")
    plt.close(fig)



def make_top20_table_view(df: pd.DataFrame, group_col: str, label_col: str) -> pd.DataFrame:
    """Create the two-column top 20 table used in the fact-sheet preview."""
    if group_col not in df.columns or "GT" not in df.columns:
        return pd.DataFrame(columns=[label_col, "GT mio"])

    work = df[[group_col, "GT"]].copy()
    work[group_col] = work[group_col].apply(normalize_country_name)
    work = work[~work[group_col].str.casefold().isin({"", "unknown", "nan", "none"})]
    work["GT"] = pd.to_numeric(work["GT"], errors="coerce").fillna(0)

    top = (
        work.groupby(group_col, as_index=False)["GT"]
        .sum()
        .sort_values(["GT", group_col], ascending=[False, True])
        .head(TOP_TABLE_N)
        .reset_index(drop=True)
    )

    top[label_col] = [f"{rank} {name}" for rank, name in enumerate(top[group_col], start=1)]
    top["GT mio"] = (top["GT"] / 1_000_000).round(2)
    return top[[label_col, "GT mio"]]


def write_top20_table_block(ws, start_row: int, start_col: int, title: str, table: pd.DataFrame) -> None:
    """Write the bordered Top 20 table exactly as a simple two-column table."""
    from openpyxl.styles import Alignment, Border, Side

    n_rows = TOP_TABLE_N + 2
    n_cols = 2
    end_row = start_row + n_rows - 1
    end_col = start_col + n_cols - 1

    thin_gray = Side(style="thin", color="D9D9D9")
    medium_black = Side(style="medium", color="000000")

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    headers = list(table.columns)
    ws.cell(row=start_row + 1, column=start_col, value=headers[0]).font = Font(bold=True)
    ws.cell(row=start_row + 1, column=start_col + 1, value=headers[1]).font = Font(bold=True)

    for idx in range(TOP_TABLE_N):
        excel_row = start_row + 2 + idx
        if idx < len(table):
            ws.cell(row=excel_row, column=start_col, value=table.iloc[idx, 0])
            ws.cell(row=excel_row, column=start_col + 1, value=float(table.iloc[idx, 1]))
            ws.cell(row=excel_row, column=start_col + 1).number_format = "0.00"
        else:
            ws.cell(row=excel_row, column=start_col, value="")
            ws.cell(row=excel_row, column=start_col + 1, value="")

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            top = medium_black if row == start_row else thin_gray
            bottom = medium_black if row == end_row else thin_gray
            left = medium_black if col == start_col else thin_gray
            right = medium_black if col == end_col else thin_gray

            if row == start_row:
                bottom = medium_black

            cell = ws.cell(row=row, column=col)
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
            cell.alignment = Alignment(
                horizontal="right" if col == start_col + 1 and row >= start_row + 2 else "left",
                vertical="center",
            )

    ws.cell(row=start_row, column=start_col).alignment = Alignment(horizontal="center", vertical="center")


def create_top20_table_sheet(preview_path: Path, df: pd.DataFrame) -> None:
    """Create the restored Top 20 table sheet with two bordered tables."""
    preview_df = prepare_preview_data(df)

    operator_table = make_top20_table_view(preview_df, "Operator_Domicile_Normalized", "Operatørland")
    flag_table = make_top20_table_view(preview_df, "Flag_Normalized", "Flag")

    wb = load_workbook(preview_path)

    if "Top 20 tabeller" in wb.sheetnames:
        del wb["Top 20 tabeller"]

    insert_idx = 1 if "Top 20" in wb.sheetnames else 0
    ws = wb.create_sheet("Top 20 tabeller", insert_idx)

    ws.sheet_view.showGridLines = True
    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 31
    ws.column_dimensions["E"].width = 13

    write_top20_table_block(ws, 1, 1, "20 største operatører", operator_table)
    write_top20_table_block(ws, 1, 4, "20 største flagstater", flag_table)

    wb.save(preview_path)


def create_top20_chart_sheet(preview_path: Path, df: pd.DataFrame) -> None:
    """Create a Top 20 sheet with chart images matching the fact-sheet style."""
    preview_df = prepare_preview_data(df)

    operator_table = make_top_chart_table(preview_df, "Operator_Domicile_Normalized", "Operatørland")
    flag_table = make_top_chart_table(preview_df, "Flag_Normalized", "Flag")

    chart_dir = preview_path.with_suffix("")
    chart_dir = chart_dir.parent / f"{chart_dir.name}_charts"
    chart_dir.mkdir(parents=True, exist_ok=True)

    op_png = chart_dir / "top_operated_fleet.png"
    flag_png = chart_dir / "top_flag_fleet.png"

    make_top_horizontal_bar_chart_png(operator_table, "Største opererede flåder i mio GT.", op_png)
    make_top_horizontal_bar_chart_png(flag_table, "Største flagstater i mio GT.", flag_png)

    wb = load_workbook(preview_path)

    if "Top 20" in wb.sheetnames:
        del wb["Top 20"]

    ws = wb.create_sheet("Top 20", 0)
    wb.active = 0
    ws.sheet_view.showGridLines = True

    if op_png.exists():
        img = ExcelImage(str(op_png))
        img.width = 620
        img.height = 360
        ws.add_image(img, "B2")

    if flag_png.exists():
        img = ExcelImage(str(flag_png))
        img.width = 620
        img.height = 360
        ws.add_image(img, "K2")

    wb.save(preview_path)


# ============================================================
# SEGMENT DISTRIBUTION TABLES AND CHARTS
# ============================================================

SEGMENT_6 = ["Bulk", "GC", "Inland", "Misc", "Offshore", "Tankers"]
SEGMENT_5 = [
    "Bulk",
    "General cargo (incl. Ferries, container and Ro-Ro)",
    "Misc (incl. Inland waterways)",
    "Offshore",
    "Tankers",
]


def make_segment_summary(df: pd.DataFrame, group_col: str, groups: list[str]) -> pd.DataFrame:
    """Return segment summary with ships, GT and shares."""
    if df.empty:
        base = pd.DataFrame({group_col: groups})
        base["Skibe"] = 0
        base["GT"] = 0
    else:
        tmp = df.copy()
        tmp[group_col] = tmp[group_col].fillna("Unknown").astype(str).str.strip()
        grouped = (
            tmp.groupby(group_col, dropna=False)
            .agg(
                Skibe=("IMO/LR/IHS No.", "count"),
                GT=("GT", "sum"),
            )
            .reset_index()
        )
        base = pd.DataFrame({group_col: groups}).merge(grouped, on=group_col, how="left")
        base["Skibe"] = base["Skibe"].fillna(0).astype(int)
        base["GT"] = pd.to_numeric(base["GT"], errors="coerce").fillna(0)

    total_ships = int(base["Skibe"].sum())
    total_gt = float(base["GT"].sum())

    base["Andel-skibe"] = base["Skibe"] / total_ships if total_ships else 0
    base["Andel-GT"] = base["GT"] / total_gt if total_gt else 0

    total_row = pd.DataFrame([{
        group_col: "Total",
        "Skibe": total_ships,
        "GT": total_gt,
        "Andel-skibe": 1 if total_ships else 0,
        "Andel-GT": 1 if total_gt else 0,
    }])

    return pd.concat([base, total_row], ignore_index=True)


def segment_summary_to_horizontal_table(summary: pd.DataFrame, group_col: str) -> pd.DataFrame:
    """Convert segment summary to fact-sheet horizontal table."""
    headers = summary[group_col].astype(str).tolist()

    table = pd.DataFrame(
        [
            ["Segment", *headers],
            ["Skibe", *summary["Skibe"].tolist()],
            ["GT", *summary["GT"].tolist()],
            ["Andel-skibe", *summary["Andel-skibe"].tolist()],
            ["Andel-GT", *summary["Andel-GT"].tolist()],
        ]
    )
    return table


def write_raw_table(ws, table: pd.DataFrame, start_row: int, start_col: int, title: str | None = None, boxed: bool = False) -> None:
    """Write a table without pandas header/index."""
    if title:
        ws.cell(row=start_row, column=start_col, value=title).font = Font(bold=True)
        start_row += 1

    for r_idx in range(table.shape[0]):
        for c_idx in range(table.shape[1]):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=table.iat[r_idx, c_idx])
            if r_idx == 0 or c_idx == 0:
                cell.font = Font(bold=True if r_idx == 0 or boxed else False)

            if isinstance(cell.value, float):
                row_label = str(table.iat[r_idx, 0])
                if "Andel" in row_label:
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "#,##0"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"

    if boxed:
        from openpyxl.styles import Border, Side
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")
        n_rows, n_cols = table.shape
        for r in range(start_row, start_row + n_rows):
            for c in range(start_col, start_col + n_cols):
                cell = ws.cell(row=r, column=c)
                cell.border = Border(
                    top=medium if r == start_row else thin,
                    bottom=medium if r == start_row + n_rows - 1 else thin,
                    left=medium if c == start_col else thin,
                    right=medium if c == start_col + n_cols - 1 else thin,
                )


def make_segment_distribution_chart_png(summary_5: pd.DataFrame, group_col: str, title: str, out_png: Path) -> None:
    """Render grouped bar chart: share of vessels vs share of GT."""
    plt, mticker = require_matplotlib()

    data = summary_5[summary_5[group_col] != "Total"].copy()
    if data.empty:
        return

    labels = data[group_col].astype(str).tolist()
    share_vessels = pd.to_numeric(data["Andel-skibe"], errors="coerce").fillna(0).tolist()
    share_gt = pd.to_numeric(data["Andel-GT"], errors="coerce").fillna(0).tolist()

    x = list(range(len(labels)))
    width = 0.34

    fig, ax = plt.subplots(figsize=(8.4, 5.0), dpi=180)

    bars1 = ax.bar([i - width / 2 for i in x], share_vessels, width=width, color=BLUE, label="Share of vessels")
    bars2 = ax.bar([i + width / 2 for i in x], share_gt, width=width, color=ORANGE, label="Share og Gross tonnage")

    ax.set_title(title, fontsize=12, fontweight="bold", color=TEXT_GREY, pad=12)
    ax.set_ylim(0, max(max(share_vessels), max(share_gt), 0.01) * 1.22)
    ax.yaxis.set_major_formatter(mticker.PercentFormatter(1.0, decimals=0))
    ax.yaxis.grid(False)
    ax.xaxis.grid(False)
    ax.set_axisbelow(True)

    wrapped_labels = [
        "Bulk",
        "General cargo (incl.\nFerries, container\nand Ro-Ro)",
        "Misc (incl. Inland\nwaterways)",
        "Offshore",
        "Tankers",
    ]

    ax.set_xticks(x)
    ax.set_xticklabels(wrapped_labels, fontsize=8, color=TEXT_GREY)
    ax.tick_params(axis="y", colors=TEXT_GREY, labelsize=8, length=0)
    ax.tick_params(axis="x", length=0)

    for spine in ax.spines.values():
        spine.set_visible(False)

    for bars, values in [(bars1, share_vessels), (bars2, share_gt)]:
        for bar, value in zip(bars, values):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + ax.get_ylim()[1] * 0.015,
                f"{value:.0%}",
                ha="center",
                va="bottom",
                fontsize=8,
                color=TEXT_GREY,
            )

    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.15), ncol=2, frameon=False, fontsize=8)
    fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def add_segment_distribution_sheet(preview_path: Path, df: pd.DataFrame) -> None:
    """Add Segmentfordeling sheet with tables and chart images."""
    preview_df = prepare_preview_data(df)

    dk_operated = preview_df[preview_df["Is_DK_Operated"]].copy()
    dk_flag = preview_df[preview_df["Is_DK_Flag"]].copy()

    op_6 = make_segment_summary(dk_operated, "Segment", SEGMENT_6)
    flag_6 = make_segment_summary(dk_flag, "Segment", SEGMENT_6)

    op_5 = make_segment_summary(dk_operated, "FactSheetGroup", SEGMENT_5)
    flag_5 = make_segment_summary(dk_flag, "FactSheetGroup", SEGMENT_5)

    op_6_table = segment_summary_to_horizontal_table(op_6, "Segment")
    flag_6_table = segment_summary_to_horizontal_table(flag_6, "Segment")
    op_5_table = segment_summary_to_horizontal_table(op_5, "FactSheetGroup")
    flag_5_table = segment_summary_to_horizontal_table(flag_5, "FactSheetGroup")

    chart_dir = preview_path.with_suffix("")
    chart_dir = chart_dir.parent / f"{chart_dir.name}_charts"
    chart_dir.mkdir(parents=True, exist_ok=True)

    op_total_ships = int(op_5.loc[op_5["FactSheetGroup"] == "Total", "Skibe"].iloc[0])
    op_total_gt_mio = round(float(op_5.loc[op_5["FactSheetGroup"] == "Total", "GT"].iloc[0]) / 1_000_000)
    flag_total_ships = int(flag_5.loc[flag_5["FactSheetGroup"] == "Total", "Skibe"].iloc[0])
    flag_total_gt_mio = round(float(flag_5.loc[flag_5["FactSheetGroup"] == "Total", "GT"].iloc[0]) / 1_000_000)

    op_title = f"Opereret flåde: {op_total_ships} skibe {op_total_gt_mio} mio. BT"
    flag_title = f"Dansk flagflåde: {flag_total_ships} skibe {flag_total_gt_mio} mio. BT"

    op_png = chart_dir / "segment_operated_distribution.png"
    flag_png = chart_dir / "segment_flag_distribution.png"

    make_segment_distribution_chart_png(op_5, "FactSheetGroup", op_title, op_png)
    make_segment_distribution_chart_png(flag_5, "FactSheetGroup", flag_title, flag_png)

    wb = load_workbook(preview_path)

    if "Segmentfordeling" in wb.sheetnames:
        del wb["Segmentfordeling"]

    insert_idx = 1 if "Top 20" in wb.sheetnames else 0
    ws = wb.create_sheet("Segmentfordeling", insert_idx)

    ws.sheet_view.showGridLines = True

    write_raw_table(ws, op_6_table, start_row=4, start_col=1, title="Operatør", boxed=True)
    write_raw_table(ws, flag_6_table, start_row=4, start_col=12, title="Flag Dansk flag", boxed=True)

    write_raw_table(ws, op_5_table, start_row=13, start_col=1, title="Operatør", boxed=False)
    write_raw_table(ws, flag_5_table, start_row=13, start_col=12, title="Flag Dansk flag", boxed=False)

    if op_png.exists():
        img = ExcelImage(str(op_png))
        img.width = 560
        img.height = 360
        ws.add_image(img, "B22")

    if flag_png.exists():
        img = ExcelImage(str(flag_png))
        img.width = 560
        img.height = 360
        ws.add_image(img, "K22")

    for col_idx in range(1, 24):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["L"].width = 16

    wb.save(preview_path)


# ============================================================
# GENERIC PREVIEW WORKBOOK
# ============================================================

def make_breakdown(df: pd.DataFrame, group_col: str, label_col: str) -> pd.DataFrame:
    """Generic breakdown by segment or ship type."""
    columns = [label_col, "Ships", "GT", "Mio. GT", "Deadweight", "TEU", "Andel af GT"]
    if df.empty or group_col not in df.columns:
        return pd.DataFrame(columns=columns)

    tmp = df.copy()
    tmp[group_col] = tmp[group_col].fillna("Unknown").astype(str).str.strip()
    tmp[group_col] = tmp[group_col].replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})

    result = (
        tmp.groupby(group_col, dropna=False)
        .agg(
            Ships=("IMO/LR/IHS No.", "count"),
            GT=("GT", "sum"),
            Deadweight=("Deadweight", "sum"),
            TEU=("TEU", "sum"),
        )
        .reset_index()
        .rename(columns={group_col: label_col})
    )

    if group_col == "Segment":
        order = {name: i for i, name in enumerate(["Bulk", "GC", "Inland", "Misc", "Offshore", "Tankers", "Unknown"])}
        result["_sort"] = result[label_col].map(order).fillna(999)
        result = result.sort_values(["_sort", label_col]).drop(columns="_sort")
    else:
        result = result.sort_values("GT", ascending=False)

    result["Mio. GT"] = (result["GT"] / 1_000_000).round(2)
    total_gt = result["GT"].sum()
    result["Andel af GT"] = result["GT"] / total_gt if total_gt else 0

    total_row = pd.DataFrame([{
        label_col: "Total",
        "Ships": result["Ships"].sum(),
        "GT": result["GT"].sum(),
        "Mio. GT": round(result["GT"].sum() / 1_000_000, 2),
        "Deadweight": result["Deadweight"].sum(),
        "TEU": result["TEU"].sum(),
        "Andel af GT": 1 if total_gt else 0,
    }])

    return pd.concat([result[columns], total_row], ignore_index=True)


def make_source_control(df: pd.DataFrame) -> pd.DataFrame:
    """Control table: files read and how they map to broad segments."""
    columns = ["Source.Name", "Segment", "Ships", "GT", "Mio. GT", "Deadweight", "TEU"]
    if "Source.Name" not in df.columns:
        return pd.DataFrame(columns=columns)

    result = (
        df.groupby(["Source.Name", "Segment"], dropna=False)
        .agg(
            Ships=("IMO/LR/IHS No.", "count"),
            GT=("GT", "sum"),
            Deadweight=("Deadweight", "sum"),
            TEU=("TEU", "sum"),
        )
        .reset_index()
        .sort_values(["Segment", "Source.Name"], key=lambda s: s.map(lambda x: natural_key(str(x))))
    )
    result["Mio. GT"] = (result["GT"] / 1_000_000).round(2)
    return result[columns]


def write_preview_section(writer, sheet_name: str, title: str, df: pd.DataFrame, startrow: int, startcol: int = 0) -> int:
    """Write one titled dataframe section and return the next available row."""
    df.to_excel(writer, sheet_name=sheet_name, startrow=startrow + 1, startcol=startcol, index=False)
    ws = writer.sheets[sheet_name]

    title_cell = ws.cell(row=startrow + 1, column=startcol + 1)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=12)

    header_row = startrow + 2
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for col_idx in range(startcol + 1, startcol + len(df.columns) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    return startrow + len(df) + 4


def format_preview_worksheet(ws) -> None:
    """Basic formatting for preview workbook."""
    ws.freeze_panes = "A2"

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        width = 14

        for row_idx in range(1, min(ws.max_row, 60) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                width = max(width, min(len(str(value)) + 2, 35))

        ws.column_dimensions[col_letter].width = width

    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, float):
                header = str(ws.cell(row=2 if row_idx > 1 else 1, column=col_idx).value or "")
                if "Andel" in header:
                    cell.number_format = "0.0%"
                elif "Mio" in header:
                    cell.number_format = "0.00"
                else:
                    cell.number_format = "#,##0.00"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"

    for row_idx in range(1, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        if any(value == "Total" for value in row_values):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)


def create_faktaark_preview(
    df: pd.DataFrame,
    preview_path: Path,
    input_files_read: int,
    rows_before_scope_filter: int,
    patrol_removed: int,
    rows_after_scope_filter: int,
    rows_before_dedup: int,
    rows_after_dedup: int,
    failed_files: list[tuple[str, str]],
) -> None:
    """Create separate Excel workbook showing the numbers that can later feed Faktaark.xlsx."""
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    if preview_path.exists():
        preview_path.unlink()

    preview_df = prepare_preview_data(df)

    top_operator = make_rank_table(preview_df, "Operator_Domicile_Normalized", "Operator Domicile")
    top_flag = make_rank_table(preview_df, "Flag_Normalized", "Flag")

    dk_operated = preview_df[preview_df["Is_DK_Operated"]].copy()
    dk_flag = preview_df[preview_df["Is_DK_Flag"]].copy()

    dk_operated_segments = make_breakdown(dk_operated, "Segment", "Segment")
    dk_flag_segments = make_breakdown(dk_flag, "Segment", "Segment")

    dk_operated_shiptypes = make_breakdown(dk_operated, "Ship Type", "Ship Type")
    dk_flag_shiptypes = make_breakdown(dk_flag, "Ship Type", "Ship Type")

    source_control = make_source_control(preview_df)

    control = pd.DataFrame(
        [
            ["Kørselsdato", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Inputmappe", str(INPUT_FOLDER)],
            ["Verdensflåde-output", str(OUT_PATH)],
            ["Preview-output", str(preview_path)],
            ["Inputfiler læst", input_files_read],
            ["Rækker før Patrol Vessel-filter", rows_before_scope_filter],
            ["Patrol Vessel-rækker fjernet", patrol_removed],
            ["Rækker efter Patrol Vessel-filter", rows_after_scope_filter],
            ["Rækker før IMO-dubletfjernelse", rows_before_dedup],
            ["IMO-dubletter fjernet", rows_before_dedup - rows_after_dedup],
            ["Endelige rækker", rows_after_dedup],
            ["Fejlede filer", len(failed_files)],
        ],
        columns=["Kontrolpunkt", "Værdi"],
    )

    failed = pd.DataFrame(failed_files, columns=["Fil", "Fejl"]) if failed_files else pd.DataFrame(columns=["Fil", "Fejl"])

    with pd.ExcelWriter(preview_path, engine="openpyxl") as writer:
        write_preview_section(writer, "Opereret_flag", "Top 15 operatørlande efter GT", top_operator, startrow=0, startcol=0)
        write_preview_section(writer, "Opereret_flag", "Top 15 flagstater efter GT", top_flag, startrow=0, startcol=10)

        next_row = write_preview_section(writer, "DK_segmenter", "Dansk opereret flåde fordelt på segment", dk_operated_segments, startrow=0, startcol=0)
        write_preview_section(writer, "DK_segmenter", "Dansk flagflåde fordelt på segment", dk_flag_segments, startrow=next_row, startcol=0)

        next_row = write_preview_section(writer, "DK_fartojstyper", "Dansk opereret flåde fordelt på Ship Type", dk_operated_shiptypes, startrow=0, startcol=0)
        write_preview_section(writer, "DK_fartojstyper", "Dansk flagflåde fordelt på Ship Type", dk_flag_shiptypes, startrow=next_row, startcol=0)

        source_control.to_excel(writer, sheet_name="Kildekontrol", index=False)
        control.to_excel(writer, sheet_name="Kontrol", index=False)
        failed.to_excel(writer, sheet_name="Fejlede_filer", index=False)

        for ws in writer.book.worksheets:
            format_preview_worksheet(ws)

    create_top20_chart_sheet(preview_path, df)
    create_top20_table_sheet(preview_path, df)
    add_segment_distribution_sheet(preview_path, df)



# ============================================================
# UPDATE EXISTING FAKTAARK TRY WORKBOOK
# ============================================================

def _set_cell_value(ws, row: int, col: int, value) -> None:
    """Set a value without changing existing formatting."""
    ws.cell(row=row, column=col).value = value


def _write_rows(ws, start_row: int, start_col: int, rows: list[list]) -> None:
    """Write a 2D list to an existing formatted range."""
    for r_idx, row_values in enumerate(rows):
        for c_idx, value in enumerate(row_values):
            _set_cell_value(ws, start_row + r_idx, start_col + c_idx, value)


def _resolve_sheet_name(wb, prefix: str) -> str:
    """Find sheet by exact name or prefix to handle truncated Excel sheet names."""
    if prefix in wb.sheetnames:
        return prefix

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith(prefix) or prefix.startswith(sheet_name):
            return sheet_name

    raise KeyError(f"Missing sheet starting with: {prefix}")


def _write_top20_into_faktaark(ws, preview_df: pd.DataFrame) -> None:
    """
    Update the existing fact-sheet sheet for:
    Opereret flåde og flagflåde.

    It writes the source tables and visible Top 20 tables while preserving formatting.
    """

    operator_rank = make_rank_table(
        preview_df,
        "Operator_Domicile_Normalized",
        "Operator Domicile",
        top_n=TOP_TABLE_N,
    )

    flag_rank = make_rank_table(
        preview_df,
        "Flag_Normalized",
        "Flag",
        top_n=TOP_TABLE_N,
    )

    operator_simple = make_top20_table_view(
        preview_df,
        "Operator_Domicile_Normalized",
        "Operatørland",
    )

    flag_simple = make_top20_table_view(
        preview_df,
        "Flag_Normalized",
        "Flag",
    )

    # Source table for operated-fleet chart/data.
    # B3:E24
    _set_cell_value(ws, 3, 2, "Operator Domicile")
    _write_rows(ws, 4, 2, [["Country ", "GT", "Rank_GT", "Label"]])

    operator_rows = []
    for _, row in operator_rank.iterrows():
        operator_rows.append([
            row["Operator Domicile"],
            float(row["GT"]),
            int(row["Rank"]),
            row["Label"],
        ])
    operator_rows += [["", "", "", ""]] * max(0, TOP_TABLE_N - len(operator_rows))
    _write_rows(ws, 5, 2, operator_rows[:TOP_TABLE_N])

    # Source table for flag-fleet chart/data.
    # H3:K24
    _set_cell_value(ws, 3, 8, "Flag")
    _write_rows(ws, 4, 8, [["Country ", "GT", "Rank_GT", "Label"]])

    flag_rows = []
    for _, row in flag_rank.iterrows():
        flag_rows.append([
            row["Flag"],
            float(row["GT"]),
            int(row["Rank"]),
            row["Label"],
        ])
    flag_rows += [["", "", "", ""]] * max(0, TOP_TABLE_N - len(flag_rows))
    _write_rows(ws, 5, 8, flag_rows[:TOP_TABLE_N])

    # Visible/simple Top 20 tables.
    # O3:P24 and S3:T24
    _set_cell_value(ws, 3, 15, "20 største operatører")
    _write_rows(ws, 4, 15, [["Operatørland", "GT mio"]])

    op_simple_rows = []
    for _, row in operator_simple.iterrows():
        op_simple_rows.append([row["Operatørland"], float(row["GT mio"])])
    op_simple_rows += [["", ""]] * max(0, TOP_TABLE_N - len(op_simple_rows))
    _write_rows(ws, 5, 15, op_simple_rows[:TOP_TABLE_N])

    _set_cell_value(ws, 3, 19, "20 største flagstater")
    _write_rows(ws, 4, 19, [["Flag", "GT mio"]])

    flag_simple_rows = []
    for _, row in flag_simple.iterrows():
        flag_simple_rows.append([row["Flag"], float(row["GT mio"])])
    flag_simple_rows += [["", ""]] * max(0, TOP_TABLE_N - len(flag_simple_rows))
    _write_rows(ws, 5, 19, flag_simple_rows[:TOP_TABLE_N])

    # Core number formats. Existing cell style is kept.
    for row in range(5, 5 + TOP_TABLE_N):
        ws.cell(row=row, column=3).number_format = "#,##0"
        ws.cell(row=row, column=9).number_format = "#,##0"
        ws.cell(row=row, column=16).number_format = "0.00"
        ws.cell(row=row, column=20).number_format = "0.00"


def _write_horizontal_fact_table(ws, start_row: int, start_col: int, title: str, table: pd.DataFrame) -> None:
    """
    Write a segment table into an existing formatted range.
    start_row/start_col points to the title cell.
    """

    _set_cell_value(ws, start_row, start_col, title)

    for r_idx in range(table.shape[0]):
        row_label = str(table.iat[r_idx, 0])
        for c_idx in range(table.shape[1]):
            cell = ws.cell(row=start_row + 1 + r_idx, column=start_col + c_idx)
            cell.value = table.iat[r_idx, c_idx]

            if row_label == "GT" and c_idx > 0:
                cell.number_format = "#,##0"
            elif "Andel" in row_label or "Share" in row_label:
                cell.number_format = "0.00%"


def _write_segments_into_faktaark(ws, preview_df: pd.DataFrame) -> None:
    """
    Update the existing fact-sheet sheet for:
    Tonnage og fordeling på fartøjs.

    It writes:
    - B4:I9    operator, 6 segments
    - K4:R9    Danish flag, 6 segments
    - B13:H18  operator, fact-sheet grouping
    - K13:Q18  Danish flag, fact-sheet grouping
    """

    dk_operated = preview_df[preview_df["Is_DK_Operated"]].copy()
    dk_flag = preview_df[preview_df["Is_DK_Flag"]].copy()

    op_6 = make_segment_summary(dk_operated, "Segment", SEGMENT_6)
    flag_6 = make_segment_summary(dk_flag, "Segment", SEGMENT_6)

    op_5 = make_segment_summary(dk_operated, "FactSheetGroup", SEGMENT_5)
    flag_5 = make_segment_summary(dk_flag, "FactSheetGroup", SEGMENT_5)

    op_6_table = segment_summary_to_horizontal_table(op_6, "Segment")
    flag_6_table = segment_summary_to_horizontal_table(flag_6, "Segment")
    op_5_table = segment_summary_to_horizontal_table(op_5, "FactSheetGroup")
    flag_5_table = segment_summary_to_horizontal_table(flag_5, "FactSheetGroup")

    _write_horizontal_fact_table(ws, 4, 2, "Operatør", op_6_table)
    _write_horizontal_fact_table(ws, 4, 11, "Flag Dansk flag", flag_6_table)

    _write_horizontal_fact_table(ws, 13, 2, "Operatør", op_5_table)
    _write_horizontal_fact_table(ws, 13, 11, "Flag Dansk flag", flag_5_table)


def update_faktaark_try(combined: pd.DataFrame, faktaark_path: Path, out_path: Path | None = None) -> None:
    """
    Write updated Top 20 and segment tables into an existing Faktaark try workbook.

    The function preserves existing formatting/charts and only updates underlying cells.
    Excel will refresh chart-linked values when the workbook is opened.
    """

    if out_path is None:
        out_path = faktaark_path

    if not faktaark_path.exists():
        raise FileNotFoundError(f"Faktaark try file not found: {faktaark_path}")

    preview_df = prepare_preview_data(combined)

    wb = load_workbook(faktaark_path)

    top20_sheet = _resolve_sheet_name(wb, "Opereret flåde og flagflåde (in")
    segment_sheet = _resolve_sheet_name(wb, "Tonnage og fordeling på fartøjs")

    _write_top20_into_faktaark(wb[top20_sheet], preview_df)
    _write_segments_into_faktaark(wb[segment_sheet], preview_df)

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)



# ============================================================
# MAIN FLOW
# ============================================================

def read_all_input_frames(file_map: list[tuple[Path, str]]) -> tuple[list[pd.DataFrame], list[tuple[str, str]]]:
    frames: list[pd.DataFrame] = []
    failed_files: list[tuple[str, str]] = []

    worker_count = max(1, min(MAX_READ_WORKERS, len(file_map)))
    print(f"Reading input files: {len(file_map)} file(s), {worker_count} worker(s)")

    if worker_count == 1:
        for file_path, source_name in file_map:
            try:
                frames.append(read_source_frame(file_path, source_name))
            except Exception as exc:
                failed_files.append((str(file_path), str(exc)))
        return frames, failed_files

    ordered_frames: list[pd.DataFrame | None] = [None] * len(file_map)
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(read_source_frame, file_path, source_name): (index, file_path)
            for index, (file_path, source_name) in enumerate(file_map)
        }

        for future in as_completed(futures):
            index, file_path = futures[future]
            try:
                ordered_frames[index] = future.result()
            except Exception as exc:
                failed_files.append((str(file_path), str(exc)))

    frames = [frame for frame in ordered_frames if frame is not None]
    return frames, failed_files


def build_world_fleet(file_map: list[tuple[Path, str]]) -> tuple[pd.DataFrame, dict[str, int], list[tuple[str, str]]]:
    frames, failed_files = read_all_input_frames(file_map)

    if not frames:
        sys.exit("No readable input files. Stopping.")

    combined = pd.concat(frames, ignore_index=True)
    combined = normalize_columns(combined)

    rows_before_scope_filter = len(combined)
    combined, patrol_removed = remove_patrol_vessels(combined)
    rows_after_scope_filter = len(combined)

    for col in OUTPUT_COLUMNS:
        if col not in combined.columns:
            combined[col] = ""

    combined = combined[OUTPUT_COLUMNS]
    combined = coerce_final_types(combined)

    rows_before_dedup = len(combined)
    if "IMO/LR/IHS No." in combined.columns:
        combined = combined.drop_duplicates(subset=["IMO/LR/IHS No."], keep="first").copy()
    rows_after_dedup = len(combined)

    stats = {
        "rows_before_scope_filter": rows_before_scope_filter,
        "patrol_removed": patrol_removed,
        "rows_after_scope_filter": rows_after_scope_filter,
        "rows_before_dedup": rows_before_dedup,
        "rows_after_dedup": rows_after_dedup,
    }

    return combined, stats, failed_files


def write_world_fleet_output(combined: pd.DataFrame) -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    if OUT_PATH.exists():
        OUT_PATH.unlink()

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME, index=False)
        format_output_worksheet(writer.sheets[OUTPUT_SHEET_NAME])


def main() -> None:
    file_map = detect_input_files(INPUT_FOLDER)
    combined, stats, failed_files = build_world_fleet(file_map)

    write_world_fleet_output(combined)

    if CREATE_FACTAARK_PREVIEW:
        create_faktaark_preview(
            combined,
            FACTAARK_PREVIEW_PATH,
            input_files_read=len(file_map),
            rows_before_scope_filter=stats["rows_before_scope_filter"],
            patrol_removed=stats["patrol_removed"],
            rows_after_scope_filter=stats["rows_after_scope_filter"],
            rows_before_dedup=stats["rows_before_dedup"],
            rows_after_dedup=stats["rows_after_dedup"],
            failed_files=failed_files,
        )

    if UPDATE_FAKTAARK_TRY:
        update_faktaark_try(
            combined=combined,
        faktaark_path=_resolve_faktaark_try_path(FAKTAARK_TRY_PATH),
            out_path=FAKTAARK_TRY_OUT_PATH,
        )

    print("Done")
    print(f"Input files read: {len(file_map)}")
    print(f"Rows before Patrol Vessel filter: {stats['rows_before_scope_filter']}")
    print(f"Patrol Vessel rows removed: {stats['patrol_removed']}")
    print(f"Rows after Patrol Vessel filter: {stats['rows_after_scope_filter']}")
    print(f"Rows before IMO deduplication: {stats['rows_before_dedup']}")
    print(f"IMO duplicates removed: {stats['rows_before_dedup'] - stats['rows_after_dedup']}")
    print(f"Final rows: {stats['rows_after_dedup']}")
    print(f"Output workbook: {OUT_PATH}")

    if CREATE_FACTAARK_PREVIEW:
        print(f"Faktaark preview workbook: {FACTAARK_PREVIEW_PATH}")

    if UPDATE_FAKTAARK_TRY:
        print(f"Faktaark try updated: {FAKTAARK_TRY_OUT_PATH}")

    if failed_files:
        print("Files skipped due to errors:")
        for file_name, error in failed_files:
            print(f"- {file_name}: {error}")


if __name__ == "__main__":
    main()
